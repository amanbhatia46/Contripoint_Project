'''
FUncitonal

'''
from datetime import datetime
import unittest
import time
import HtmlTestRunner as x
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook


class Css_Certifications():

    def __init__(self):
        self.filename = '/home/am.bhatia/Desktop/contripoint/abc1.xlsx'
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'certicss' not in self.wb.sheetnames:
            self.wb.create_sheet('certicss')
            self.wb.save('/home/am.bhatia/Desktop/contripoint/abc1.xlsx')
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'certicss':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.ws['A1'] = 'Module Name'
        self.ws['B1'] = 'URL css'
        self.ws['C1'] = 'figma css'
        self.ws['E1'] = 'Result'
        self.ws['F1'] = 'Date and Time'
        self.wb.save('/home/am.bhatia/Desktop/contripoint/abc1.xlsx')

    def writeWS(self, row, cell1D, cell2D, cell3D, cell4D):
        self.ws[f'A{row}'] = cell1D
        self.ws[f'B{row}'] = cell2D
        self.ws[f'C{row}'] = cell3D
        self.ws[f'E{row}'] = cell4D
        self.ws[f'F{row}'] = str(datetime.now())
        self.wb.save('/home/am.bhatia/Desktop/contripoint/abc1.xlsx')

    def setUp(self):
        '''
          The setUp() methods allows to define set of instructions.If this method it self raises an exception while executing tests, the test methods will not be executed.
        '''
        s = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=s)

        print("\n \n\n\n\n>>>>>>>>>>CSS PROPERTIES >>>>>>>>>>>>")

        driver = self.driver
        driver.maximize_window()
        driver.get(
            "http://dev-contripoint.geminisolutions.com.s3-website.ap-south-1.amazonaws.com/#/dashboard")
        button = driver.find_element(
            By.XPATH, '/html/body/app-root/div/div/app-login-screen/div/div/div/mat-card/div[2]/div/button')
        button.click()
        time.sleep(3)

        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[1])
        driver.find_element(
            By.XPATH, '//*[@id="identifierId"]').send_keys('test.user@geminisolutions.in')
        driver.find_element(
            By.XPATH, '//*[@id="identifierNext"]/div/button/span').click()
        time.sleep(3)

        driver.find_element(
            By.XPATH, '//*[@id="i0118"]').send_keys(login_Password)
        driver.find_element(
            By.XPATH, '//*[@id="passwordNext"]/div/button/span').click()
        time.sleep(6)

        driver.switch_to.window(window_handles[0])
        time.sleep(6)
        print("\n  1 - Gemini Id and Password successfully login")

    def test_Certifications_readFontProperty(self):
        try:
            Certifications = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-dash-board/div[3]/app-dash-cards/div/div[1]/div[1]/mat-card/mat-card-header/div[2]/mat-card-title')
            print("\n Certifications:- ")
            print("Certifications font-size: " +
                  Certifications.value_of_css_property("font-size") + "\n  FIGMA font-size: 16px")
            pix = "\n FIGMA font-size: 16px".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "font-size") == pix else 'fail'
            print(result)
            self.writeWS(2, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "font-size"), "FIGMA font-size: 16px", str(result))

            print("\n Certifications color: " + Certifications.value_of_css_property(
                "color") + "\n  FIGMA Color: rgba(0, 0, 0, 1)")
            pix = "\n FIGMA Color: rgba(0, 0, 0, 1)".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "color") == pix else 'fail'
            print(result)
            self.writeWS(3, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "color"), "FIGMA Color: rgba(0, 0, 0, 1)", (result))

            print("\n Certifications font-family: " + Certifications.value_of_css_property(
                "font-family") + "\n  FIGMA font-family: Poppins")
            pix = "\n FIGMA font-family: Poppins".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "font-family") == pix else 'fail'
            print(result)
            self.writeWS(4, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "font-family"), "FIGMA font-family: Poppins", (result))

            print("\n Certifications text-align: " + Certifications.value_of_css_property(
                "text-align") + "\n FIGMA text-align: left")
            pix = "\n FIGMA text-align: left".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "text-align") == pix else 'fail'
            print(result)
            self.writeWS(5, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "text-align"), "FIGMA text-align: left", (result))

            print("\n Certifications font-style: " + Certifications.value_of_css_property(
                "font-style") + "\n FIGMA font-style: normal")
            pix = "\n FIGMA font-style: normal".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "font-style") == pix else 'fail'
            print(result)
            self.writeWS(6, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "font-style"), "FIGMA font-style: normal", (result))

            print("\n Certifications font-weight: " + Certifications.value_of_css_property(
                "font-weight") + "\n  FIGMA font-weight: 600")
            pix = "\n FIGMA font-weight: 600".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "font-weight") == pix else 'fail'
            print(result)
            self.writeWS(7, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "font-weight"), "FIGMA font-weight: 600", str(result))

            print("\n Certifications line-height: " + Certifications.value_of_css_property(
                "line-height") + "\n  FIGMA  line-height: 24px")
            pix = "\n FIGMA  line-height: 24px".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "line-height") == pix else 'fail'
            print(result)
            self.writeWS(8, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "line-height"), "FIGMA  line-height: 24px", str(result))

            print("\n Certifications letter-spacing: " + Certifications.value_of_css_property(
                "letter-spacing") + "\n FIGMA  letter-spacing: 0.03em")
            pix = "\n FIGMA  letter-spacing: 0.03em".split(': ')[1]
            result = 'pass' if Certifications.value_of_css_property(
                "letter-spacing") == pix else 'fail'
            print(result)
            self.writeWS(9, 'CERTIFICATIONS', Certifications.value_of_css_property(
                "letter-spacing"), "FIGMA  letter-spacing: 0.03em", str(result))

            print("\n Certifications Location : " +
                  str(Certifications.location))

            print("\n Certifications Size : " + str(Certifications.size))

            Certifications.click()
            time.sleep(10)

            NameOfCertification = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-certificate-table/div/div/div/div[2]/div/table/thead/tr/th[1]/div/div[1]')
            print("\n  NameOfCertification:- ")
            print("NameOfCertification font-size: " +
                  NameOfCertification.value_of_css_property("font-size") + "\n FIGMA font-size: 14px")
            pix = "\n FIGMA font-size: 14px".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "font-size") == pix else 'fail'
            print(result)
            self.writeWS(12, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "font-size"), "FIGMA font-size: 14px", str(result))

            print("\n NameOfCertification color: " + NameOfCertification.value_of_css_property(
                "color") + "\n FIGMA color: rgba(0, 0, 0, 1)")
            pix = "\n FIGMA Color: rgba(0, 0, 0, 1)".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "color") == pix else 'fail'
            print(result)
            self.writeWS(13, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "color"), "FIGMA Color: rgba(0, 0, 0, 1)", (result))

            print("\n NameOfCertification font-family: " +
                  NameOfCertification.value_of_css_property("font-family") + "\n FIGMA font-family: Poppins")
            pix = "\n FIGMA font-family: Poppins".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "font-family") == pix else 'fail'
            print(result)
            self.writeWS(14, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "font-family"), "FIGMA font-family: Poppins", (result))

            print("\n NameOfCertification font-style: " +
                  NameOfCertification.value_of_css_property("font-style") + "\n FIGMA font-style: normal")
            pix = "\n FIGMA font-style: normal".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "font-style") == pix else 'fail'
            print(result)
            self.writeWS(15, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "font-style"), "FIGMA font-style: normal", (result))

            print("\n NameOfCertification text-align: " +
                  NameOfCertification.value_of_css_property("text-align") + "\n FIGMA text-align: left")
            pix = "\n FIGMA text-align: left".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "text-align") == pix else 'fail'
            print(result)
            self.writeWS(16, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "text-align"), "FIGMA text-align: left", (result))

            print("\n NameOfCertification font-weight: " +
                  NameOfCertification.value_of_css_property("font-weight") + "\n FIGMA font-weight: 600")
            pix = "\n FIGMA font-weight: 600".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "font-weight") == pix else 'fail'
            print(result)
            self.writeWS(17, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "font-weight"), "FIGMA font-weight: 600", str(result))

            print("\n NameOfCertification line-height: " +
                  NameOfCertification.value_of_css_property("line-height") + "\n FIGMA line-height: 21px")
            pix = "\n FIGMA  line-height: 21px".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "line-height") == pix else 'fail'
            print(result)
            self.writeWS(18, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "line-height"), "FIGMA  line-height: 21px", str(result))

            print("\n NameOfCertification letter-spacing: " + NameOfCertification.value_of_css_property(
                "letter-spacing") + "\n FIGMA letter-spacing: 0.36px")
            pix = "\n FIGMA  letter-spacing: 0.36px".split(': ')[1]
            result = 'pass' if NameOfCertification.value_of_css_property(
                "letter-spacing") == pix else 'fail'
            print(result)
            self.writeWS(19, 'Name Of Certification', NameOfCertification.value_of_css_property(
                "letter-spacing"), "FIGMA  letter-spacing: 0.36px", str(result))

            print("\n Location : " + str(NameOfCertification.location))

            print("\n Size : " + str(NameOfCertification.size))
            time.sleep(6)

            CompletionDate = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-certificate-table/div/div/div/div[2]/div/table/thead/tr/th[2]/div/div[1]')
            print("\n  Completion Date:- ")
            print("Completion Date font-size: " +
                  CompletionDate.value_of_css_property("font-size") + "\n FIGMA font-size: 14px")
            pix = "\n FIGMA font-size: 14px".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "font-size") == pix else 'fail'
            print(result)
            self.writeWS(21, 'CompletionDate', CompletionDate.value_of_css_property(
                "font-size"), "FIGMA font-size: 14px", str(result))

            print("\n Completion Date color: " + CompletionDate.value_of_css_property(
                "color") + "\n FIGMA color: rgba(0, 0, 0, 1)")
            pix = "\n FIGMA Color: rgba(0, 0, 0, 1)".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "color") == pix else 'fail'
            print(result)
            self.writeWS(22, 'CompletionDate', CompletionDate.value_of_css_property(
                "color"), "FIGMA Color: rgba(0, 0, 0, 1)", (result))

            print("\n Completion Date font-family: " + CompletionDate.value_of_css_property(
                "font-family") + "\n FIGMA font-family: Poppins")
            pix = "\n FIGMA font-family: Poppins".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "font-family") == pix else 'fail'
            print(result)
            self.writeWS(23, 'CompletionDate', CompletionDate.value_of_css_property(
                "font-family"), "FIGMA font-family: Poppins", (result))

            print("\n Completion Date font-style: " + CompletionDate.value_of_css_property(
                "font-style") + "\n FIGMA font-style: normal")
            pix = "\n FIGMA font-style: normal".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "font-style") == pix else 'fail'
            print(result)
            self.writeWS(24, 'CompletionDate', CompletionDate.value_of_css_property(
                "font-style"), "FIGMA font-style: normal", (result))

            print("\n Completion Date text-align: " +
                  CompletionDate.value_of_css_property("text-align") + "\n FIGMA text-align: left")
            pix = "\n FIGMA text-align: left".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "text-align") == pix else 'fail'
            print(result)
            self.writeWS(25, 'CompletionDate', CompletionDate.value_of_css_property(
                "text-align"), "FIGMA text-align: left", (result))

            print("\n Completion Date font-weight: " +
                  CompletionDate.value_of_css_property("font-weight") + "\n FIGMA font-weight: 600")
            pix = "\n FIGMA font-weight: 600".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "font-weight") == pix else 'fail'
            print(result)
            self.writeWS(26, 'CompletionDate', CompletionDate.value_of_css_property(
                "font-weight"), "FIGMA font-weight: 600", str(result))

            print("\n Completion Date line-height: " + CompletionDate.value_of_css_property(
                "line-height") + "\n FIGMA line-height: 21px")
            pix = "\n FIGMA  line-height: 21px".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "line-height") == pix else 'fail'
            print(result)
            self.writeWS(27, 'CompletionDate', CompletionDate.value_of_css_property(
                "line-height"), "FIGMA  line-height: 21px", str(result))

            print("\n Completion Date letter-spacing: " + CompletionDate.value_of_css_property(
                "letter-spacing") + "\n FIGMA letter-spacing: 0.36px")
            pix = "\n FIGMA  letter-spacing: 0.36px".split(': ')[1]
            result = 'pass' if CompletionDate.value_of_css_property(
                "letter-spacing") == pix else 'fail'
            print(result)
            self.writeWS(28, 'CompletionDate', CompletionDate.value_of_css_property(
                "letter-spacing"), "FIGMA  letter-spacing: 0.36px", str(result))

            print("\n Location : " + str(CompletionDate.location))
            print("\n Size : " + str(CompletionDate.size))
            time.sleep(6)

            Status = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-certificate-table/div/div/div/div[2]/div/table/thead/tr/th[3]/div/div[1]')
            print("\n  Status:- ")
            print("Status font-size: " +
                  Status.value_of_css_property("font-size") + "\n FIGMA font-size: 14px")
            pix = "\n FIGMA font-size: 14px".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "font-size") == pix else 'fail'
            print(result)
            self.writeWS(30, 'Status', Status.value_of_css_property(
                "font-size"), "FIGMA font-size: 14px", str(result))

            print("\n Status color: " + Status.value_of_css_property("color") +
                  "\n FIGMA color: rgba(0, 0, 0, 1)")
            pix = "\n FIGMA Color: rgba(0, 0, 0, 1)".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "color") == pix else 'fail'
            print(result)
            self.writeWS(31, 'Status', Status.value_of_css_property(
                "color"), "FIGMA Color: rgba(0, 0, 0, 1)", (result))

            print("\n Status font-family: " + Status.value_of_css_property(
                "font-family") + "\n FIGMA font-family: Poppins")
            pix = "\n FIGMA font-family: Poppins".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "font-family") == pix else 'fail'
            print(result)
            self.writeWS(32, 'Status', Status.value_of_css_property(
                "font-family"), "FIGMA font-family: Poppins", (result))

            print("\n Status font-style: " +
                  Status.value_of_css_property("font-style") + "\n FIGMA font-style: normal")
            pix = "\n FIGMA font-style: normal".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "font-style") == pix else 'fail'
            print(result)
            self.writeWS(33, 'Status', Status.value_of_css_property(
                "font-style"), "FIGMA font-style: normal", (result))

            print("\n Status text-align: " +
                  Status.value_of_css_property("text-align") + "\n FIGMA text-align: left")
            pix = "\n FIGMA text-align: left".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "text-align") == pix else 'fail'
            print(result)
            self.writeWS(34, 'Status', Status.value_of_css_property(
                "text-align"), "FIGMA text-align: left", (result))

            print("\n Status font-weight: " +
                  Status.value_of_css_property("font-weight") + "\n FIGMA font-weight: 600")
            pix = "\n FIGMA font-weight: 600".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "font-weight") == pix else 'fail'
            print(result)
            self.writeWS(35, 'Status', Status.value_of_css_property(
                "font-weight"), "FIGMA font-weight: 600", str(result))

            print("\n Status line-height: " +
                  Status.value_of_css_property("line-height") + "\n FIGMA line-height: 21px")
            pix = "\n FIGMA  line-height: 21px".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "line-height") == pix else 'fail'
            print(result)
            self.writeWS(36, 'Status', Status.value_of_css_property(
                "line-height"), "FIGMA  line-height: 21px", str(result))

            print("\n Status letter-spacing: " + Status.value_of_css_property(
                "letter-spacing") + "\n FIGMA letter-spacing: 0.36px")
            pix = "\n FIGMA  letter-spacing: 0.36px".split(': ')[1]
            result = 'pass' if Status.value_of_css_property(
                "letter-spacing") == pix else 'fail'
            print(result)
            self.writeWS(37, 'Status', Status.value_of_css_property(
                "letter-spacing"), "FIGMA  letter-spacing: 0.36px", str(result))

            print("\n Location : " + str(Status.location))
            print("\n Size : " + str(Status.size))
            time.sleep(6)

            Action = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-certificate-table/div/div/div/div[2]/div/table/thead/tr/th[3]/div/div[1]')
            print("\n  Action:- ")
            print("\Action font-size: " +
                  Action.value_of_css_property("font-size") + "\n FIGMA font-size: 14px")
            pix = "\n FIGMA font-size: 14px".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "font-size") == pix else 'fail'
            print(result)
            self.writeWS(39, 'Action', Action.value_of_css_property(
                "font-size"), "FIGMA font-size: 14px", str(result))

            print("\n Action color: " + Action.value_of_css_property("color") +
                  "\n FIGMA color: rgba(0, 0, 0, 1)")
            pix = "\n FIGMA Color: rgba(0, 0, 0, 1)".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "color") == pix else 'fail'
            print(result)
            self.writeWS(40, 'Action', Action.value_of_css_property(
                "color"), "FIGMA Color: rgba(0, 0, 0, 1)", (result))

            print("\n Action font-family: " + Action.value_of_css_property(
                "font-family") + "\n FIGMA font-family: Poppins")
            pix = "\n FIGMA font-family: Poppins".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "font-family") == pix else 'fail'
            print(result)
            self.writeWS(41, 'Action', Action.value_of_css_property(
                "font-family"), "FIGMA font-family: Poppins", (result))

            print("\n Action font-style: " +
                  Action.value_of_css_property("font-style") + "\n FIGMA font-style: normal")
            pix = "\n FIGMA font-style: normal".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "font-style") == pix else 'fail'
            print(result)
            self.writeWS(42, 'Action', Action.value_of_css_property(
                "font-style"), "FIGMA font-style: normal", (result))

            print("\n Action text-align: " +
                  Action.value_of_css_property("text-align") + "\n FIGMA text-align: left")
            pix = "\n FIGMA text-align: left".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "text-align") == pix else 'fail'
            print(result)
            self.writeWS(43, 'Action', Action.value_of_css_property(
                "text-align"), "FIGMA text-align: left", (result))

            print("\n Action font-weight: " +
                  Action.value_of_css_property("font-weight") + "\n FIGMA font-weight: 600")
            pix = "\n FIGMA font-weight: 600".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "font-weight") == pix else 'fail'
            print(result)
            self.writeWS(44, 'Action', Action.value_of_css_property(
                "font-weight"), "FIGMA font-weight: 600", str(result))

            print("\n Action line-height: " +
                  Action.value_of_css_property("line-height") + "\n FIGMA line-height: 21px")
            pix = "\n FIGMA  line-height: 21px".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "line-height") == pix else 'fail'
            print(result)
            self.writeWS(45, 'Action', Action.value_of_css_property(
                "line-height"), "FIGMA  line-height: 21px", str(result))

            print("\n Action letter-spacing: " + Action.value_of_css_property(
                "letter-spacing") + "\n FIGMA letter-spacing: 0.36px")
            pix = "\n FIGMA  letter-spacing: 0.03em".split(': ')[1]
            result = 'pass' if Action.value_of_css_property(
                "letter-spacing") == pix else 'fail'
            print(result)
            self.writeWS(46, 'Action', Action.value_of_css_property(
                "letter-spacing"), "FIGMA  letter-spacing: 0.03em", str(result))

            print("\n Location : " + str(Action.location))
            print("\n Size : " + str(Action.size))
            time.sleep(6)

            CreatedOn = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-certificate-table/div/div/div/div[2]/div/table/thead/tr/th[3]/div/div[1]')
            print("\n  CreatedOn:- ")
            print("Created On font-size: " +
                  CreatedOn.value_of_css_property("font-size") + "\n FIGMA font-size: 14px")
            pix = "\n FIGMA font-size: 14px".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "font-size") == pix else 'fail'
            print(result)
            self.writeWS(48, 'CreatedOn', CreatedOn.value_of_css_property(
                "font-size"), "FIGMA font-size: 14px", str(result))

            print("\n Created On color: " + CreatedOn.value_of_css_property("color") +
                  "\n FIGMA color: rgba(0, 0, 0, 1)")
            pix = "\n FIGMA Color: rgba(0, 0, 0, 1)".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "color") == pix else 'fail'
            print(result)
            self.writeWS(49, 'CreatedOn', CreatedOn.value_of_css_property(
                "color"), "FIGMA Color: rgba(0, 0, 0, 1)", (result))

            print("\n Created On font-family: " + CreatedOn.value_of_css_property(
                "font-family") + "\n FIGMA font-family: Poppins")
            pix = "\n FIGMA font-family: Poppins".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "font-family") == pix else 'fail'
            print(result)
            self.writeWS(50, 'CreatedOn', CreatedOn.value_of_css_property(
                "font-family"), "FIGMA font-family: Poppins", (result))

            print("\n Created On font-style: " +
                  CreatedOn.value_of_css_property("font-style") + "\n FIGMA font-style: normal")
            pix = "\n FIGMA font-style: normal".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "font-style") == pix else 'fail'
            print(result)
            self.writeWS(51, 'CreatedOn', CreatedOn.value_of_css_property(
                "font-style"), "FIGMA font-style: normal", (result))

            print("\n Created On text-align: " +
                  CreatedOn.value_of_css_property("text-align") + "\n FIGMA text-align: left")
            pix = "\n FIGMA text-align: left".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "text-align") == pix else 'fail'
            print(result)
            self.writeWS(52, 'CreatedOn', CreatedOn.value_of_css_property(
                "text-align"), "FIGMA text-align: left", (result))

            print("\n Created On font-weight: " +
                  CreatedOn.value_of_css_property("font-weight") + "\n FIGMA font-weight: 600")
            pix = "\n FIGMA font-weight: 600".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "font-weight") == pix else 'fail'
            print(result)
            self.writeWS(53, 'CreatedOn', CreatedOn.value_of_css_property(
                "font-weight"), "FIGMA font-weight: 600", str(result))

            print("\n Created On line-height: " +
                  CreatedOn.value_of_css_property("line-height") + "\n FIGMA line-height: 21px")
            pix = "\n FIGMA  line-height: 21px".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "line-height") == pix else 'fail'
            print(result)
            self.writeWS(54, 'CreatedOn', CreatedOn.value_of_css_property(
                "line-height"), "FIGMA  line-height: 21px", str(result))

            print("\n Created On letter-spacing: " + CreatedOn.value_of_css_property(
                "letter-spacing") + "\n FIGMA letter-spacing: 0.36px")
            pix = "\n FIGMA  letter-spacing: 0.03em".split(': ')[1]
            result = 'pass' if CreatedOn.value_of_css_property(
                "letter-spacing") == pix else 'fail'
            print(result)
            self.writeWS(55, 'CreatedOn', CreatedOn.value_of_css_property(
                "letter-spacing"), "FIGMA  letter-spacing: 0.03em", str(result))

            print("\n Location : " + str(CreatedOn.location))
            print("\n Size : " + str(CreatedOn.size))
            time.sleep(6)

            Certificates = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-certificate-table/div/div/div/div[2]/div/div[1]/div/div[1]/ul/li/h2')
            print("\n  Certificates:- ")
            print("Certificates font-size: " +
                  Certificates.value_of_css_property("font-size") + "\n FIGMA font-size: 18px")
            pix = "\n FIGMA font-size: 18px".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "font-size") == pix else 'fail'
            print(result)
            self.writeWS(57, 'Certificates', Certificates.value_of_css_property(
                "font-size"), "FIGMA font-size: 18px", str(result))

            print("\n Certificates color: " + Certificates.value_of_css_property(
                "color") + "\n FIGMA color: rgba(0, 0, 0, 1)")
            pix = "\n FIGMA Color: rgba(0, 0, 0, 1)".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "color") == pix else 'fail'
            print(result)
            self.writeWS(58, 'Certificates', Certificates.value_of_css_property(
                "color"), "FIGMA Color: rgba(0, 0, 0, 1)", (result))

            print("\n Certificates font-family: " + Certificates.value_of_css_property(
                "font-family") + "\n FIGMA font-family: Poppins")
            pix = "\n FIGMA font-family: Poppins".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "font-family") == pix else 'fail'
            print(result)
            self.writeWS(59, 'Certificates', Certificates.value_of_css_property(
                "font-family"), "FIGMA font-family: Poppins", (result))

            print("\n Certificates font-style: " + Certificates.value_of_css_property(
                "font-style") + "\n FIGMA font-style: normal")
            pix = "\n FIGMA font-style: normal".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "font-style") == pix else 'fail'
            print(result)
            self.writeWS(60, 'Certificates', Certificates.value_of_css_property(
                "font-style"), "FIGMA font-style: normal", (result))

            print("\n Certificates text-align: " +
                  Certificates.value_of_css_property("text-align") + "\n FIGMA text-align: left")
            pix = "\n FIGMA text-align: left".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "text-align") == pix else 'fail'
            print(result)
            self.writeWS(61, 'Certificates', Certificates.value_of_css_property(
                "text-align"), "FIGMA text-align: left", (result))

            print("\n Certificates font-weight: " +
                  Certificates.value_of_css_property("font-weight") + "\n FIGMA font-weight: 600")
            pix = "\n FIGMA font-weight: 600".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "font-weight") == pix else 'fail'
            print(result)
            self.writeWS(62, 'Certificates', Certificates.value_of_css_property(
                "font-weight"), "FIGMA font-weight: 600", str(result))

            print("\n Certificates line-height: " + Certificates.value_of_css_property(
                "line-height") + "\n FIGMA line-height: 27px")
            pix = "\n FIGMA  line-height: 21px".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "line-height") == pix else 'fail'
            print(result)
            self.writeWS(63, 'Certificates', Certificates.value_of_css_property(
                "line-height"), "FIGMA  line-height: 21px", str(result))

            print("\n Certificates letter-spacing: " + Certificates.value_of_css_property(
                "letter-spacing") + "\n FIGMA letter-spacing: 0.36px")
            pix = "\n FIGMA  letter-spacing: 0.36px".split(': ')[1]
            result = 'pass' if Certificates.value_of_css_property(
                "letter-spacing") == pix else 'fail'
            print(result)
            self.writeWS(64, 'Certificates', Certificates.value_of_css_property(
                "letter-spacing"), "FIGMA  letter-spacing: 0.36px", str(result))

            print("\n Location : " + str(Certificates.location))
            print("\n Size : " + str(Certificates.size))
            time.sleep(6)
            


        except Exception as e:
            print("\n \nERROR While getting CSS PROPERTIES >>>>>>>>>>>>>>>\n\n")
            print(e)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = Css_Certifications()
    tb.setUp()
    tb.test_Certifications_readFontProperty()
    # unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))
