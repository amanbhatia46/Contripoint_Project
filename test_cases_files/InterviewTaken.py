'''
FUncitonal

'''
from json import load
from telnetlib import EC

from turtle import position
import unittest
import time
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime, timedelta

from datainputs import *


class ITAuth(unittest.TestCase):

    def __init__(self):
        """ interact with the underlying operating system."""
        import os
        print(os.path.join(os.getcwd() + "\\xyz.xlsx"), ">>>>>>")
        self.filename = os.path.join(os.getcwd() + "\\xyz.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'Interview Taken' not in self.wb.sheetnames:
            self.wb.create_sheet('Interview Taken')
            self.wb.save(os.path.join(os.getcwd() + "\\xyz.xlsx"))
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'Interview Taken':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd() + "\\xyz.xlsx"))

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>> INTERVIEW TAKEN >>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            driver.implicitly_wait(10)  # seconds
            button = driver.find_element(
                By.XPATH, '//button[contains(., "Login with Gemini mail")]')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="i0116"]').send_keys(login_Id)
            self.ws['A2'] = 'aman.bhatia@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="idSIButton9"]').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="i0118"]').send_keys(login_Password)
            self.ws['B2'] = 'RitaNandini96'
            driver.find_element(
                By.XPATH, '//*[@id="idSIButton9"]').click()
            time.sleep(6)

            # Entering Multi-factor Authentication (MFA) Code Manually
            driver.find_element(By.ID, "idTxtBx_SAOTCC_OTC")
            time.sleep(6)
            print("\n Entering and Verifying MFA Code \n")

            driver.find_element(
                By.XPATH, '//*[@id="idSubmit_SAOTCC_Continue"]').click()
            time.sleep(5)

            driver.find_element(
                By.XPATH, '//*[@id="KmsiCheckboxField"]').click()
            time.sleep(5)

            driver.find_element(By.ID, "idSIButton9").click()

            print("\n Login Successfull \n")

            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

    def Interview_taken(self):
        """

        """
        try:

            self.driver.find_element(
                By.XPATH, '//*[contains(text(),"Interview Taken")]').click()
            print("\n 2 - Selecting 'Interview Taken' successfully")
            self.ws['C3'] = 'Interview Taken'
            self.ws['G3'] = 'Pass'
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Interview Taken >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 2 - Selecting Interview Taken gets fail")
            self.ws['C3'] = 'Interview Taken'
            self.ws['G3'] = 'Fail'

            self.wb.save(self.filename)

            print(
                "\n 3 - sample test case of entering 'Interview Taken' successfully open")

    def Add_New_Button(self):
        """

        """
        try:
            self.driver.execute_script("window.scrollTo(0,100)")

            b = self.driver.find_element(
                By.XPATH, "//*[@id='add_btn']").click()
            time.sleep(6)
            print("\n 4 - 'Add New' button gets selected")
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Adding New Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 4 - Add New button gets fail")
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Fail'

            self.wb.save(self.filename)

    def Interview_position(self):
        """

        """
        try:
            # # position for which Interview was taken
            c = self.driver.find_element(
                By.XPATH, '//*[@id="mat-select-0"]/div/div[2]').click()
            print("\n 5 - position done")

            print("\n 6 - Selecting Positions from the drop down list . . .")

            BussinessAnalystwithCRM = self.driver.find_element(
                By.XPATH, '//*[@id="mat-option-14"]/span').click()
            print("\n 7 - Bussiness Analyst with CRM")
            time.sleep(3)
            CloudEngineer = self.driver.find_element(
                By.XPATH, '//*[@id="mat-option-16"]/span').click()
            print("\n 8 - Cloud Engineer")
            time.sleep(3)
            CollaborationPlatform = self.driver.find_element(
                By.XPATH, '//*[@id="mat-option-18"]/span').click()
            print("\n 9 - Collaboration Platform")
            time.sleep(3)
            AdministratorDevOpsEngineer = self.driver.find_element(
                By.XPATH, '//*[@id="mat-option-22"]/span').click()
            print("\n 10 - Administrator DevOps Engineer")
            time.sleep(6)

            self.ws['C5'] = 'Interview Position'
            self.ws['G5'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Interview Position >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 5 - Interview Position gets fail")
            self.ws['C5'] = 'Interview Position'
            self.ws['G5'] = 'Fail'

            self.wb.save(self.filename)

    def ExperienceLevel(self):
        """

        """
        try:
            ExpLevel = self.driver.find_element(
                By.CSS_SELECTOR, ".cdk-overlay-transparent-backdrop").click()
            self.driver.find_element(
                By.XPATH, '//*[contains(text(),"Select Interviewee Experience Level")]').click()
            print("\n 11 - Selecting 'Interviewee Exp Level'")
            time.sleep(6)

            self.driver.find_element(
                By.XPATH, '//*[@id="mat-radio-4"]').click()
            print("\n 12 - Junior (Less than 5 years)")

            self.ws['C6'] = 'Experience Level'
            self.ws['G6'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Experience Level >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 11 - Selecting Experience Level gets fail")
            self.ws['C6'] = 'Experience Level'
            self.ws['G6'] = 'Fail'

            self.wb.save(self.filename)

    def No_of_InterviewTaken(self):
        """

        """
        try:
            IntNumber = self.driver.find_element(
                By.XPATH, '//input[@id="mat-input-0"]').send_keys("2")
            print("\n 13 - Adding No. of Interview Taken")
            time.sleep(6)

            self.ws['C7'] = 'No. of Interview Taken'
            self.ws['G7'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN No. of Interview Taken >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 13 - No. of Interview Taken gets fail")
            self.ws['C7'] = 'No. of Interview Taken'
            self.ws['G7'] = 'Fail'

            self.wb.save(self.filename)

    def Interview_Month(self):
        """

        """
        try:
            IntMonth = self.driver.find_element(
                By.XPATH, '//*[@id="mat-select-4"]/div/div[2]/div').click()
            print("\n 14 -Selecting the Month . . . ")
            time.sleep(6)

            Month = ActionChains(self.driver).move_to_element(self.driver.find_element(
                By.XPATH, '//span[text()=" January "]')).click().perform()
            print("\n 15 - Selecting 'JANUARY' Month from the list")
            time.sleep(6)

            self.ws['C8'] = 'Interview Month'
            self.ws['G8'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Interview Month >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 13 - Selecting Interview Month gets fail")
            self.ws['C8'] = 'Interview Month'
            self.ws['G8'] = 'Fail'

            self.wb.save(self.filename)

    def Goal_Type(self):
        """

        """
        try:
            # Goal Type-
            self.driver.find_element(
                By.XPATH, '//*[text()="Engineering Council (EC)"]').click()
            print("\n 14 - Selecting Goal Type- Engineering Council (EC)")
            self.ws['C9'] = 'Engineering Council (EC)'
            self.ws['G9'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN Goal Type >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 14 - Unable to select Goal_Type")
            self.ws['C9'] = 'Engineering Council (EC)'
            self.ws['G9'] = 'Fail'

            self.wb.save(self.filename)


    def Description(self):
        """

        """
        try:
            self.driver.find_element(
                By.XPATH,'//textarea[@id="mat-input-1"]').send_keys("Automation testing")
            time.sleep(4)
            print("\n 15 - Entering Description Done")
            self.ws['C10'] = 'Description'
            self.ws['G10'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Project Status >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 15 - Entering Description gets fail")
            self.ws['C10'] = 'Description'
            self.ws['G10'] = 'Fail'
            self.wb.save(self.filename)

    def Submit(self):
        """

        """
        try:
            Submit = self.driver.find_element(
                By.XPATH, '//button[contains(text(),"SUBMIT")]')
            Submit.click()
            print("\n 16 - All Details get submit successfully.")
            time.sleep(6)

            self.ws['C11'] = 'Submit'
            self.ws['G11'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN SUBMIT  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 16 - Unable to Click Submit Button")
            self.ws['C11'] = 'Submit'
            self.ws['G11'] = 'Fail'

            self.wb.save(self.filename)
            time.sleep(3)

    def OK_Button(self):
        """

        """
        try:
            OK = self.driver.find_element(
                By.XPATH, '//span[text()="OK"]').click()
            print("\n 17 - Clicking OK Button")
            self.ws['C12'] = 'OK Button'
            self.ws['G12'] = 'Pass'
            time.sleep(4)
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN OK Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 17 - Unable to Click OK Button")
            self.ws['C12'] = 'OK Button'
            self.ws['G12'] = 'Fail'

            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = ITAuth()
    tb.setUp()
    tb.Interview_taken()
    tb.Add_New_Button()
    tb.Interview_position()
    tb.ExperienceLevel()
    tb.No_of_InterviewTaken()
    tb.Interview_Month()
    tb.Goal_Type()
    tb.Description()
    tb.Submit()
    tb.OK_Button()

    # unittest.main()
#    unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))
