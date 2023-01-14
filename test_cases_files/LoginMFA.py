from lib2to3.pgen2 import driver
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from multiprocessing import Event
from traceback import print_exc
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta

from datainputs import *

class Login(unittest.TestCase):

    def __init__(self):
        
        """ interact with the underlying operating system."""
        import os
        print(os.path.join(os.getcwd()+ "\\xyz.xlsx"),">>>>>>")
        self.filename = os.path.join(os.getcwd()+ "\\xyz.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'LoginMFA' not in self.wb.sheetnames:
            self.wb.create_sheet('LoginMFA')
            self.wb.save('/C:Users/Aman Bhatia/OneDrive - Gemini Solutions/Desktop/Contripoint_Project/xyz.xlsx')
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'LoginMFA':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active  
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        self.ws['H2'] = futuredate
        
    def setExternalDriver(self):
        s = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=s)

        self.driver.maximize_window()
        return self.driver

    def MFA(self):
        try:
            # set implicit wait time
            self.driver.implicitly_wait(10)  # seconds

            print("\n >>>>>>>>>> Login >>>>>>>>>>>> \n")

            # get url
            self.driver.get(
                "https://dev-contripoint.geminisolutions.com/#/dashboard")
            button = self.driver.find_element(
                By.XPATH, '//button[contains(., "Login with Gemini mail")]')
            button.click()

            print("\n 1- Login With Gemini mail \n")

            time.sleep(6)

            window_handles = self.driver.window_handles
            self.driver.switch_to.window(window_handles[1])

            # Entering Mail Id in input
            self.driver.find_element(
                By.XPATH, '//*[@id="i0116"]').send_keys(login_Id)
            self.ws['A2'] = 'aman.bhatia@geminisolutions.com'
            print("\n 2- Entering mail id \n")

            self.driver.find_element(
                By.XPATH, '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div[1]/div[3]/div/div/div/div[4]/div/div/div/div[2]/input').click()
            time.sleep(3)

            # Entering Passowrd in Input
            self.driver.find_element(
                By.XPATH, '//*[@id="i0118"]').send_keys(login_Password)
            self.ws['B2'] = 'RitaNandini96'
            print("\n 3- Entering mail password \n")

            self.driver.find_element(
                By.XPATH, '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[4]/div[2]/div/div/div/div/input').click()
            time.sleep(6)

            # Entering Multi-factor Authentication (MFA) Code Manually
            self.driver.find_element(By.ID, "idTxtBx_SAOTCC_OTC")
            time.sleep(6)
            print("\n 4- Entering and Verifying MFA Code \n")

            self.driver.find_element(
                By.XPATH, '//*[@id="idSubmit_SAOTCC_Continue"]').click()
            time.sleep(5)

            self.driver.find_element(
                By.XPATH, '//*[@id="KmsiCheckboxField"]').click()
            time.sleep(5)

            self.driver.find_element(By.ID, "idSIButton9").click()

            print("\n 5- Login Successfull \n")

            print("\n 6 - Gemini Id and Password successfully login \n")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:

            print("\n 6 - Gemini Id and Password  login failed")
            print(e)

            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = Login()
    tb.setExternalDriver()
    tb.MFA()
