from lib2to3.pgen2 import driver
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from multiprocessing import Event
from traceback import print_exc
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta

from datainputs import *
from LoginMFA import *


class PEvent(unittest.TestCase):

    def __init__(self):
        
        """ interact with the underlying operating system."""
        import os
        print(os.path.join(os.getcwd()+ "\\xyz.xlsx"),">>>>>>")
        self.filename = os.path.join(os.getcwd()+ "\\xyz.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'Polling Event' not in self.wb.sheetnames:
            self.wb.create_sheet('Polling Event')
            self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'Polling Event':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setExternalDriver(self,driver):
        self.driver = driver

    def Events(self):
        """

        """
        try:
            time.sleep(10)
            y = self.driver.window_handles[0]
            time.sleep(6)
            self.driver.switch_to.window(y)

            # Clicking on Events
            self.driver.find_element(
                By.XPATH, '/html/body/app-root/div/app-navbar/div/div/div[1]/div/div[2]/div[1]/div[2]').click()
            time.sleep(6)
            print("\n 7 - Clicking on Events ")

            # Create New Event
            self.driver.find_element(By.ID, "add_btn").click()
            time.sleep(5)
            print("\n 8 - Creating New Event")

        except Exception as e:
            print("\n 7 - Clicking on Events ")
            print("\n 8 - Error in Creating New Event")
            print(e)

    def Add_Banner_Image(self):
        """

        """
        try:
            # Adding Banner Image
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[1]/form/div[1]/div[1]/label/img').click()
            time.sleep(8)
            pyautogui.write(banner_image)
            pyautogui.press('enter')
            print("\n 9 - Banner Image Upload successfully")
            self.ws['C3'] = 'Uploading Banner Image'
            self.ws['G3'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(3)

        except Exception as e:
            print("\n ERROR IN Adding Banner Image >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 9 - Unable to Add Banner Image")
            self.ws['C3'] = 'Uploading Banner Image'
            self.ws['G3'] = 'Fail'

            self.wb.save(self.filename)

    def Add_Listing_Image(self):
        """

        """
        try:
            # Adding Listing Image
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[1]/form/div[2]/div[1]/label/label/img').click()
            time.sleep(8)
            pyautogui.write(listing_image)
            pyautogui.press('enter')
            time.sleep(3)

            print("\n 10 - Listing Image Upload successfully")

            self.ws['C4'] = 'Upload Listing Image'
            self.ws['G4'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Adding Listing Image >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 10 - Unable to Add Listing Image")

            self.ws['C4'] = 'Upload Listing Image'
            self.ws['G4'] = 'Fail'
            self.wb.save(self.filename)

    def Event_Name(self):
        """

        """
        try:
            # Entering  Event Name
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[1]/form/div[2]/div[2]/div[1]/mat-form-field/div/div[1]/div/input').send_keys("Polling Event")
            time.sleep(5)
            print("\n 11 - Entering Event Name ")

            self.ws['C5'] = 'Event Name'
            self.ws['G5'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Event Name >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 11 - Unable to Add Event Name")

            self.ws['C5'] = 'Event Name'
            self.ws['G5'] = 'Fail'
            self.wb.save(self.filename)

    def Description(self):
        """

        """
        try:
            # Entering Description
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[1]/form/div[2]/div[2]/div[2]/mat-form-field/div/div[1]/div/textarea').send_keys("Please take participate in Polling Event and win Rewards")
            time.sleep(5)
            print("\n 12 - Entering Description  ")

            self.ws['C6'] = 'Description '
            self.ws['G6'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Description  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 12 - Unable to Add Description ")

            self.ws['C6'] = 'Description '
            self.ws['G6'] = 'Fail'
            self.wb.save(self.filename)

    def Next_Button(self):
        """

        """
        try:
            # Clicking on Next Button
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[1]/form/div[3]/div/button').click()
            time.sleep(5)
            print("\n 13 - Clicking on Next Button  ")

            self.ws['C7'] = 'Next Button '
            self.ws['G7'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Next Button  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 13 - Unable to Add Next Button ")

            self.ws['C7'] = 'Next Button '
            self.ws['G7'] = 'Fail'
            self.wb.save(self.filename)

    def Event_Type(self):
        """

        """
        try:
            # Clicking on Event Type
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[2]/form/div[1]/div/mat-form-field/div/div[1]/div/mat-select/div/div[1]/span').click()
            time.sleep(5)
            print("\n 14 - Clicking on Event Type Dropdown ")

            self.driver.find_element(By.ID, "mat-option-2").click()
            time.sleep(5)

            print("\n 15 - Event Type - Polling Event")

            self.ws['C8'] = 'Event Type - Polling Event'
            self.ws['G8'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Event Type  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 15 - Unable to Select Event Type ")

            self.ws['C8'] = 'Event Type - Polling Event'
            self.ws['G8'] = 'Fail'
            self.wb.save(self.filename)

    def Start_Date(self):
        """

        """
        try:
            # Clicking on Start Date
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[2]/form/div[2]/div[1]/mat-form-field/div/div[1]/div[1]/input').click()
            time.sleep(5)
            print("\n 16 - Clicking on Start Date Calander ")

            # self.driver.find_element(By.XPATH,'//*[@id="mat-datepicker-0"]/div/mat-month-view/table/tbody/tr[2]/td[2]/div[1]').click()
            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-calendar-body-today").click()
            time.sleep(5)

            self.ws['C9'] = 'Start Date'
            self.ws['G9'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Start Date  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 16 - Unable to Add Start Date ")

            self.ws['C9'] = 'Start Date'
            self.ws['G9'] = 'Fail'
            self.wb.save(self.filename)

    def End_Date(self):
        """

        """
        try:
            # Clicking on End Date
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[2]/form/div[2]/div[2]/mat-form-field/div/div[1]/div[1]/input').click()
            time.sleep(5)
            print("\n 17 - Clicking on End Date Calander ")

            # self.driver.find_element(By.XPATH,'//*[@id="mat-datepicker-0"]/div/mat-month-view/table/tbody/tr[2]/td[2]/div[1]').click()
            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-calendar-body-today").click()
            time.sleep(5)

            self.ws['C10'] = 'End Date'
            self.ws['G10'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN End Date  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 17 - Unable to Add End Date ")

            self.ws['C10'] = 'End Date'
            self.ws['G10'] = 'Fail'
            self.wb.save(self.filename)

    def Reward(self):
        """

        """
        try:
            # Entering Reward
            self.driver.find_element(
                By.ID, "mat-input-4").send_keys("Cash Rs1000/-")
            time.sleep(5)
            print("\n 18 - Entering Reward  ")

            self.ws['C11'] = 'Reward '
            self.ws['G11'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Reward  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 18 - Unable to Add Reward ")

            self.ws['C11'] = 'Reward '
            self.ws['G11'] = 'Fail'
            self.wb.save(self.filename)

    def Next_Button1(self):
        """

        """
        try:
            # Clicking on Next Button
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[2]/form/div[4]/div[2]/button').click()
            time.sleep(5)
            print("\n 19 - Clicking on Next Button  ")

            self.ws['C12'] = 'Next Button '
            self.ws['G12'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Next Button  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 19 - Unable to Add Next Button ")

            self.ws['C12'] = 'Next Button '
            self.ws['G12'] = 'Fail'
            self.wb.save(self.filename)

    def Voters_Category(self):
        """
        """
        try:
            # Clicking on Voters Category
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[3]/form/div[1]/div[1]/mat-form-field/div/div[1]/div/mat-select/div/div[1]/span').click()
            time.sleep(5)
            print("\n 20 - Clicking on Voters Category  ")

            # Selecting Individual Participants From Dropdown
            self.driver.find_element(
                By.ID, "mat-option-5").click()
            time.sleep(5)
            print("\n 21 - Selecting Individual Participants")

            self.ws['C13'] = 'Voters Category - Individual Participants '
            self.ws['G13'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Voters Category  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 21 - Unable To Select Individual Participants")

            self.ws['C13'] = 'Voters Category - Individual Participants '
            self.ws['G13'] = 'Fail'
            self.wb.save(self.filename)

    def Eligible_Voters(self):
        """

        """
        try:
            # Inserting Voter's Name (Aman Bhatia) in Field
            self.driver.find_element(
                By.ID, "mat-input-6").send_keys('Aman Bhatia')
            time.sleep(5)

            # Clicking on pseudo-checkbox
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[3]/form/div[1]/div[2]/mat-form-field/div/div[1]/div[3]/mat-selection-list/mat-list-option/div/mat-pseudo-checkbox').click()
            time.sleep(5)
            print("\n 22 - Eligible Voter- Aman Bhatia")

            # clear the text entered in Field
            self.driver.find_element(By.ID, "mat-input-6").clear()

            # Inserting Voter's Name (Alpana Upadhyay) in Field
            self.driver.find_element(
                By.ID, "mat-input-6").send_keys('Alpana Upadhyay')
            time.sleep(5)

            # Clicking on pseudo-checkbox
            self.driver.find_element(
                By.XPATH, '//*[@id="cdk-step-content-0-2"]/form/div[1]/div[2]/mat-form-field/div/div[1]/div[3]/mat-selection-list/mat-list-option/div/mat-pseudo-checkbox').click()
            time.sleep(5)
            print("\n 23 - Eligible Voter- Alpana Upadhyay")

            print("\n 24 - Eligible Voters - Aman Bhatia, Alpana Upadhyay")

            self.ws['C14'] = 'Eligible Voters - Aman Bhatia, Alpana Upadhyay'
            self.ws['G14'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Eligible Voters  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print(
                "\n 24 - Unable To Select Eligible Voters - Aman Bhatia, Alpana Upadhyay")

            self.ws['C14'] = 'Eligible Voters - Aman Bhatia, Alpana Upadhyay'
            self.ws['G14'] = 'Fail'
            self.wb.save(self.filename)

    def Next_Button2(self):
        """

        """
        try:
            # Clicking on Next Button
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[3]/form/div[2]/div[2]/button').click()
            time.sleep(5)
            print("\n 25 - Clicking on Next Button  ")

            self.ws['C15'] = 'Next Button '
            self.ws['G15'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Next Button  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 25 - Unable to Add Next Button ")

            self.ws['C15'] = 'Next Button '
            self.ws['G15'] = 'Fail'
            self.wb.save(self.filename)

    def Participant_Category(self):
        """
        """
        try:
            # Clicking on Participant Category
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[4]/form/div[1]/div[1]/mat-form-field/div/div[1]/div/mat-select/div/div[1]/span').click()
            time.sleep(5)
            print("\n 26 - Clicking on Participant Category  ")

            # Selecting Individual Participants From Dropdown
            self.driver.find_element(
                By.ID, "mat-option-3").click()
            time.sleep(5)
            print("\n 27 - Selecting Individual Participants")

            self.ws['C16'] = 'Participant Category - Individual Participants '
            self.ws['G16'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Participant Category  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 27 - Unable To Select Individual Participants")

            self.ws['C16'] = 'Participant Category - Individual Participants '
            self.ws['G16'] = 'Fail'
            self.wb.save(self.filename)

    def Eligible_Participant(self):
        """

        """
        try:
            # Inserting Voter's Name (Aman Bhatia) in Field
            self.driver.find_element(
                By.ID, "mat-input-5").send_keys('Aman Bhatia')
            time.sleep(5)

            # Clicking on pseudo-checkbox
            self.driver.find_element(
                By.XPATH, '//*[@id="cdk-step-content-0-3"]/form/div[1]/div[2]/mat-form-field/div/div[1]/div[3]/mat-selection-list/mat-list-option/div/mat-pseudo-checkbox').click()
            time.sleep(5)
            print("\n 28 - Eligible Voter- Aman Bhatia")

            # clear the text entered in Field
            self.driver.find_element(By.ID, "mat-input-5").clear()

            # Inserting Voter's Name (Alpana Upadhyay) in Field
            self.driver.find_element(
                By.ID, "mat-input-5").send_keys('Alpana Upadhyay')
            time.sleep(5)

            # Clicking on pseudo-checkbox
            self.driver.find_element(
                By.XPATH, '//*[@id="cdk-step-content-0-3"]/form/div[1]/div[2]/mat-form-field/div/div[1]/div[3]/mat-selection-list/mat-list-option/div/mat-pseudo-checkbox').click()
            time.sleep(5)
            print("\n 29 - Eligible Voter- Alpana Upadhyay")

            print("\n 30 - Eligible Participant - Aman Bhatia, Alpana Upadhyay")

            self.ws['C17'] = 'Eligible Participant - Aman Bhatia, Alpana Upadhyay'
            self.ws['G17'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Eligible Participant  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 30 - Unable To Select Eligible Participant - Aman Bhatia, Alpana Upadhyay")

            self.ws['C17'] = 'Eligible Participant - Aman Bhatia, Alpana Upadhyay'
            self.ws['G17'] = 'Fail'
            self.wb.save(self.filename)

    def Submit(self):
        """

        """
        try:
            # Clicking on Submit
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-event-modal/mat-dialog-content/mat-horizontal-stepper/div[2]/div[4]/form/div[2]/div[2]/button').click()
            time.sleep(5)
            print("\n 31 - Clicking on Submit ")

            self.ws['C18'] = 'Submit '
            self.ws['G18'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n ERROR IN Submit  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 31 - Unable to Add Submit ")

            self.ws['C18'] = 'Submit '
            self.ws['G18'] = 'Fail'
            self.wb.save(self.filename)

    def OK_Button(self):
        """

        """
        try:
            # OK Button

            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-confirmation-modal/div/div[4]/button').click()
            print("\n 32 - Clicking OK Button")
            self.ws['C19'] = 'OK Button'
            self.ws['G19'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN OK Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 32 - Unable to Click OK Button")
            self.ws['C19'] = 'OK Button'
            self.ws['G19'] = 'Fail'

            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = PEvent()
    login = Login()
    dr = login.setExternalDriver()
    login.MFA()
    tb.setExternalDriver(driver=dr)
    tb.Events()
    tb.Add_Banner_Image()
    tb.Add_Listing_Image()
    tb.Event_Name()
    tb.Description()
    tb.Next_Button()
    tb.Event_Type()
    tb.Start_Date()
    tb.End_Date()
    tb.Reward()
    tb.Next_Button1()
    tb.Voters_Category()
    tb.Eligible_Voters()
    tb.Next_Button2()
    tb.Participant_Category()
    tb.Eligible_Participant()
    tb.Submit()
    tb.OK_Button()
