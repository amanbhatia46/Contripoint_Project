from typing import KeysView
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from multiprocessing import Event
from traceback import print_exc
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta

from datainputs import *
from LoginMFA import *



class RBugEvent(unittest.TestCase):

    def __init__(self):
        self.filename = '/home/am.bhatia/Desktop/contripoint/ABC1.xlsx'
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'Report Bug' not in self.wb.sheetnames:
            self.wb.create_sheet('Report Bug')
            self.wb.save('/home/am.bhatia/Desktop/contripoint/ABC1.xlsx')
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'Report Bug':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save('/home/am.bhatia/Desktop/contripoint/ABC1.xlsx')

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setExternalDriver(self,driver):
        self.driver = driver

    def Report_Bug(self):
        """
            Report Bug
        """
        try:

            window_handles = self.driver.window_handles
            self.driver.switch_to.window(window_handles[0])
            time.sleep(5)

            # Report a Bug/Feedback
            Bug = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div/app-footer/div/div/div/div/div[2]/div/div[1]/button')
            time.sleep(6)
            Bug.click()

            print("\n 7 - Selecting 'Report a Bug/Feedback' . .")
            time.sleep(6)

            self.ws['C3'] = 'Report a Bug/Feedback'
            self.ws['G3'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\n ERROR IN Report a Bug/Feedback >>>>>>>>>>>>>>>\n\n")
            print("\n 7 - Selecting Report a Bug/Feedback gets fail")

            self.ws['C3'] = 'Report a Bug/Feedback'
            self.ws['G3'] = 'Fail'
            self.wb.save(self.filename)

    def Add_New(self):
        """

        """
        try:
            self.driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-bug-report-table/div/div/div/div[2]/div/div/div/div[2]/ul/div/button').send_keys(KeysView.CONTROL + Keys.HOME)

            self.driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-bug-report-table/div/div/div/div[2]/div/div/div/div[2]/ul/div/button').click()
            time.sleep(5)
            print("\n 8 - Clickong on  ADD NEW button")
            self.ws['C4'] = 'Add New'
            self.ws['G4'] = 'Pass'

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\n ERROR IN ADD NEW >>>>>>>>>>>>>>>\n\n")
            print("\n 8 - Clickong on Add New Button gets fail")
            self.ws['C4'] = 'Add New'
            self.ws['G4'] = 'Fail'

            self.wb.save(self.filename)

    def Issue_Subject(self):
        """

        """
        try:
            self.driver.find_element(
                By.ID, "mat-input-25").send_keys("Design work")
            time.sleep(5)
            print("\n 9 - Issue Subject - 'Design work' ")
            self.ws['C5'] = 'Issue Subject'
            self.ws['G5'] = 'Pass'

            self.wb.save(self.filename)


        except Exception as e:
            print("\n\n ERROR IN ISSUE SUBJECT >>>>>>>>>>>>>>>\n\n")
            print("\n 9 - Selecting Issue Subject gets fail")
            self.ws['C5'] = 'Issue Subject'
            self.ws['G5'] = 'Fail'

            self.wb.save(self.filename)

    def Issue_Category(self):
        """

        """
        try:
            Dropdown = self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-bug-report-modal/div/mat-dialog-content/form/div/div[2]/div[1]/mat-form-field/div/div[1]/div')
            time.sleep(5)
            Dropdown.click()
            time.sleep(5)

            IC = self.driver.find_element(
                By.ID, "mat-option-75").click()
            print("\n 10 - Selecting 'Design'")
            time.sleep(5)

            self.ws['C6'] = 'Issue Category - Design'
            self.ws['G6'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\n ERROR IN Selecting Issue Category >>>>>>>>>>>>>>>\n\n")
            print("\n 10 - Selecting 'Issue Category' gets fail")
            self.ws['C6'] = 'Issue Category - Design'
            self.ws['G6'] = 'Fail'

            self.wb.save(self.filename)

    def Select_Website_Feature(self):
        """

        """
        try:
            Dropdown = self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-bug-report-modal/div/mat-dialog-content/form/div/div[2]/div[2]/mat-form-field/div/div[1]/div')
            time.sleep(5)
            Dropdown.click()
            time.sleep(5)

            SWF = self.driver.find_element(
                By.ID, "mat-option-78").click()
            print("\n 11 - Selecting 'Certificate' as Website Feature")
            time.sleep(5)

            self.ws['C7'] = 'Website Feature - Certificate'
            self.ws['G7'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\n ERROR IN Selecting Website Feature >>>>>>>>>>>>>>>\n\n")
            print("\n 11 - Selecting 'Website Feature' gets fail")
            self.ws['C7'] = 'Website Feature - Certificate'
            self.ws['G7'] = 'Fail'

            self.wb.save(self.filename)

    def Description(self):
        """
            Enter Event's Description
        """
        try:
            self.driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-bug-report-modal/div/mat-dialog-content/form/div/div[4]/div/mat-form-field/div/div[1]/div').send_keys(
                "Reporting a designing bug to developers.")

            print("\n 12 - Entering Description Done")
            time.sleep(5)

            self.ws['C8'] = 'Description'
            self.ws['G8'] = 'Pass'

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\n ERROR IN DESCRIPTION >>>>>>>>>>>>>>>\n\n")
            print("\n 12 - Entering Description gets fail")

            self.ws['C8'] = 'Description'
            self.ws['G8'] = 'Fail'

            self.wb.save(self.filename)

    def attachment_button(self):
        """
            Add Listing Image
        """
        try:

            Add_Attachment = self.driver.find_element(
                By.ID, "attachment_btn")
            Add_Attachment.click()
            print("\n 13 - Clicking on Attachment Button")
            time.sleep(8)

            pyautogui.write('/home/am.bhatia/Pictures/Wallpapers/tree.jpeg')
            time.sleep(8)
            pyautogui.press('enter')
            time.sleep(5)
            print("\n 14 - Attachment Upload successfully")

            self.ws['C9'] = 'Uploading Attachment gets Pass'
            self.ws['G9'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\n ERROR IN ATTACHMENT >>>>>>>>>>>>>>>\n\n")
            print("\n 14 - Attachment Uploading gets fail")
            self.ws['C9'] = 'Uploading Attachment gets Fail'
            self.ws['G9'] = 'Fail'

            self.wb.save(self.filename)

    def Submit(self):
        '''
        '''
        try:
            submit = self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-bug-report-modal/div/div[2]/div[1]/button')
            time.sleep(5)
            submit.click()

            print("\n 15 - All Bug Details get SUBMIT successfully.")

            self.ws['C10'] = 'Submit'
            self.ws['G10'] = 'Pass'
            self.wb.save(self.filename)

            element = self.driver.find_element_by_tag_name('h2')

            if element.text != u'Design work':
                print("Verify Passed : element text is %r") % element.text
            else:
                print("Verify Failed : element text is not %r") % element.text

        except Exception as e:
            print("\n\n ERROR IN SUBMIT >>>>>>>>>>>>>>>\n\n")
            print("\n 15 - Unable to submit all details")

            self.ws['C10'] = 'Submit'
            self.ws['G10'] = 'Fail'
            self.wb.save(self.filename)

    def OK_Button(self):
        '''
            OK Button
        '''
        try:

            ok = self.driver.find_element(By.ID, "ok_btn")
            time.sleep(5)
            ok.click()

            print("\n 16 - Clicking OK Button")
            time.sleep(5)

            self.ws['C11'] = 'OK Button'
            self.ws['G11'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\n ERROR IN OKButton >>>>>>>>>>>>>>>\n\n")
            print("\n 16 - Unable to Click OK Button")

            self.ws['C11'] = 'OK Button'
            self.ws['G11'] = 'Fail'
            self.wb.save(self.filename)

    def Logout(self):
        """
        """
        try:
            Icon = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div/app-navbar/div/div/div[2]/div/div[2]').click()
            print("\n 17 - Clicking on icon . .")

            time.sleep(5)

            logout = self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/div/div/button[2]').click()
            print("\n 18 - Logout Successfull")

            self.ws['C12'] = 'Logout'
            self.ws['G12'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\n ERROR IN LOGOUT >>>>>>>>>>>>>>>\n\n")
            print("\n 18 - Unable to Logout")

            self.ws['C12'] = 'Logout'
            self.ws['G12'] = 'Fail'
            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = RBugEvent()
    login = Login()
    dr = login.setExternalDriver()
    login.MFA()
    tb.setExternalDriver(driver=dr)
    tb.Report_Bug()
    tb.Add_New()
    tb.Issue_Subject()
    tb.Issue_Category()
    tb.Select_Website_Feature()
    tb.Description()
    tb.attachment_button()
    tb.Submit()
    tb.OK_Button()
    tb.Logout()
