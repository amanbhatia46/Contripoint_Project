# from json import load
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta


class TSAuth(unittest.TestCase):

    def __init__(self):
        
        """ interact with the underlying operating system."""
        import os
        print(os.path.join(os.getcwd()+ "\\xyz.xlsx"),">>>>>>")
        self.filename = os.path.join(os.getcwd()+ "\\xyz.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'TrainingSessions' not in self.wb.sheetnames:
            self.wb.create_sheet('TrainingSessions')
            self.wb.save('/C:Users/Aman Bhatia/OneDrive - Gemini Solutions/Desktop/Contripoint_Project/xyz.xlsx')
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'TrainingSessions':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))


        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>> PROJECTS >>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            button = driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-login-screen/div/div/div/mat-card/div[2]/div/button')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="identifierId"]').send_keys('test.user@geminisolutions.in')
            self.ws['A2'] = 'test.user@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="identifierNext"]/div/button/span').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input').send_keys('Gemini@123#')
            self.ws['B2'] = 'Gemini@123#'
            driver.find_element(
                By.XPATH, '//*[@id="passwordNext"]/div/button/span').click()
            time.sleep(6)

            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

    def Training_Session(self):
        """

        """
        try:
            
            # Training&Session
            trainingsession = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-dash-board/div[3]/app-dash-cards/div/div[1]/div[3]/mat-card/mat-card-header/div[2]/mat-card-title')
            trainingsession.click()
            print("\n 2 - Selecting 'TRAINING & SESSION'")
            time.sleep(20)
            self.ws['C3'] = 'TRAINING & SESSION'
            self.ws['G3'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN TRAINING & SESSION >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 2 - Selecting TRAINING & SESSION gets fail")
            self.ws['C3'] = 'TRAINING & SESSION'
            self.ws['G3'] = 'Fail'

            self.wb.save(self.filename)

    def Add_New(self):
        """

        """
        try:
            # ADD NEW
            AddNew = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-training-session-table/div/div/div/div[2]/div/div/div/div[2]/ul/div/button').click()
            print("\n 3 - 'Add New' button gets selected")
            time.sleep(10)
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Pass'

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Adding New Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 3 - Add New button gets fail")
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Fail'

            self.wb.save(self.filename)

    def SessionName(self):
        """

        """
        try:
            # sessionName
            assert self.driver.find_element(
                By.CSS_SELECTOR, ".ng-tns-c85-16 > .mat-form-field-infix").text == "Name of Training & Session Conducted *"
            self.driver.find_element(By.ID, "mat-input-0").click()
            self.driver.find_element(
                By.ID, "mat-input-0").send_keys("Selenium with Python")
            print("\n 4 - Name of Training & Session Conducted - 'Selenium with Python'")
            self.ws['C5'] = 'SessionName'
            self.ws['G5'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

            # Sessions
            self.driver.find_element(
                By.XPATH, '//*[@id="mat-input-1"]').send_keys("2")
            print("\n 5 - Sessions Done ")
            time.sleep(5)
            

        except Exception as e:
            print("\n\nERROR IN SESSION NAME >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 4 - Adding SessionName gets fail")
            self.ws['C5'] = 'SessionName'
            self.ws['G5'] = 'Fail'

            self.wb.save(self.filename)

    def Description(self):
        """

        """
        try:
            # Enter Description
            self.driver.find_element(
                By.ID, "mat-input-2").send_keys("very good")
            print("\n 6 - Description")
            self.ws['C6'] = 'Description'
            self.ws['G6'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN Project Status >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 6 - Entering Description gets fail")
            self.ws['C6'] = 'Description'
            self.ws['G6'] = 'Fail'
            self.wb.save(self.filename)

    def NumberOfSessionProvided(self):
        """

        """
        try:
            # No. Of Training&Session Provided
            self.driver.find_element(By.ID, "mat-input-3").send_keys("5")
            print("\n 7 - No. Of Training & Session Provided")
            time.sleep(5)
            self.ws['C7'] = 'Number Of Session Provided'
            self.ws['G7'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Number Of Session Provided >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 7 - Number Of Session Provided gets fail")
            self.ws['C7'] = 'Number Of Session Provided'
            self.ws['G7'] = 'Fail'
            self.wb.save(self.filename)

    def DateOfSession(self):
        '''
        '''
        try:
            # Date Of Session
            actions = self.driver.find_element(
                By.XPATH, '//*[@id="mat-dialog-0"]/app-training-modal/div/mat-dialog-content/form/div[4]/div[2]/mat-form-field/div/div[1]/div[2]/mat-datepicker-toggle/button')
            actions.click()
            print("\n 8 - Opening Calander to select the date . . .")
            self.ws['C8'] = 'Date of Session'
            self.ws['G8'] = 'Pass'
            self.wb.save(self.filename)

            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[4]/div/mat-datepicker-content/div[2]/mat-calendar/div/mat-month-view/table/tbody/tr[3]/td[4]').click()
            time.sleep(5)
            print("\n 9 - Date Of Session Done")
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN Date OF Session >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 8 - Date of Session gets fail")
            self.ws['C8'] = 'Date of Session'
            self.ws['G8'] = 'Fail'
            self.wb.save(self.filename)

    def Technology(self):
        """
        """
        try:
            # Technology Discussed In Training&Session
            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-select-placeholder").click()
            print("\n 10 - Opening Technology List . . .")
            self.ws['C9'] = 'Technology'
            self.ws['G9'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(6)

            self.driver.find_element(
                By.XPATH, '//*[@id="mat-option-3"]').click()
            print("\n 11 - 'Automation Testing' gets select from the list")
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN Technology >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 10 -'Technology' gets fail")
            self.ws['C9'] = 'Technology'
            self.ws['G9'] = 'Fail'
            self.wb.save(self.filename)

    def Submit(self):
        """

        """
        try:
            # Submit
            Submit = ActionChains(self.driver).move_to_element(self.driver.find_element(
                By.XPATH, '//*[@id="add_btn"]'))
            Submit.click().perform()
            time.sleep(5)
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-training-modal/div/div[2]/div[1]/button').click()
            print("\n 12 - All Details get submit successfully.")
            time.sleep(5)
            self.ws['C10'] = 'Submit'
            self.ws['G10'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN SUBMIT  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 12 - Unable to Click Submit Button")
            self.ws['C10'] = 'Submit'
            self.ws['G10'] = 'Fail'

            self.wb.save(self.filename)
            time.sleep(3)

    def OK_Button(self):
        """

        """
        try:
            # OK Button
            assert self.driver.find_element(
                By.CSS_SELECTOR, "#ok_btn > .mat-button-wrapper").text == "OK"
            time.sleep(5)

            element = self.driver.find_element(By.ID, "ok_btn")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            time.sleep(5)

            self.driver.find_element(By.ID, "ok_btn").click()
            print("\n 13 - Clicking OK Button")
            self.ws['C11'] = 'OK Button'
            self.ws['G11'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN OK Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 13 - Unable to Click OK Button")
            self.ws['C11'] = 'OK Button'
            self.ws['G11'] = 'Fail'

            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = TSAuth()
    tb.setUp()
    tb.Training_Session()
    tb.Add_New()
    tb.SessionName()
    tb.Description()
    tb.NumberOfSessionProvided()
    tb.DateOfSession()
    tb.Technology()
    tb.Submit()
    tb.OK_Button()

    # unittest.main()
#    unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))
