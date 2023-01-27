'''
FUncitonal

'''
import requests
from json import load
import unittest
import time
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from xlutils.copy import copy as xl_copy
from datetime import datetime, timedelta

from datainputs import *

class Api(unittest.TestCase):

    def __init__(self):
        self.filename = os.path.join(os.getcwd()+ "\\result.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        print(self.wb)
        print(self.wb.sheetnames)

        # self.ws.append([self.filename])
        self.wb.create_sheet('Certifications')
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd()+ "\\result.xlsx"))

        # datetime object containing current date and time
        futuredate = datetime.now()
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)

            self.driver.maximize_window()
            print("\n\n\n\n\n>>>>>>>>> CERTIFICATIONS >>>>>>>>>>>>")

            # self.driver = self.self.driver
            # set implicit wait time
            self.driver.implicitly_wait(10)  # seconds

            print("\n >>>>>>>>>> Login >>>>>>>>>>>> \n")

            # get url
            self.driver.get(
                "https://dev-contripoint.geminisolutions.com/#/dashboard")
            button = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-login-screen/div/div/div/mat-card/div[2]/div/button')
            button.click()

            print("\n 1- Login With Gemini mail \n")

            time.sleep(6)

            window_handles = self.driver.window_handles
            self.driver.switch_to.window(window_handles[1])

            # Entering Mail Id in input
            self.driver.find_element(
                By.XPATH, '//*[@id="i0116"]').send_keys(login_Id)
            self.ws['A2'] = 'aman.bhatia@geminisolutions.com'
            print("\n 2- Entering mail id \n")

            self.driver.find_element(
                By.XPATH, '//input[@id='idSIButton9']').click()
            time.sleep(3)

            # Entering Passowrd in Input
            self.driver.find_element(
                By.XPATH, '//*[@id="i0118"]').send_keys(login_Password)
            self.ws['B2'] = 'RitaNandini96'
            print("\n 3- Entering mail password \n")

            self.driver.find_element(
                By.XPATH, '/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[4]/div[2]/div/div/div/div/input').click()
            time.sleep(6)

            # Entering Multi-factor Authentication (MFA) Code Manually
            self.driver.find_element(By.ID, "idTxtBx_SAOTCC_OTC")
            time.sleep(6)
            print("\n 4- Entering and Verifying MFA Code \n")

            self.driver.find_element(
                By.XPATH, '//*[@id="idSubmit_SAOTCC_Continue"]').click()
            time.sleep(5)

            self.driver.find_element(
                By.XPATH, '//*[@id="KmsiCheckboxField"]').click()
            time.sleep(5)

            self.driver.find_element(By.ID, "idSIButton9").click()

            print("\n 5- Login Successfull \n")

            print("\n 6 - Gemini Id and Password successfully login \n")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:

            print("\n 6 - Gemini Id and Password  login failed")
            print(e)

            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

     # As per unittest module, individual test should start with test_
    def Certifications(self):
        """

        """
        try:
            #Certifications
            a = self.driver.find_element(By.XPATH,'/html/body/app-root/div[3]/div/app-dash-board/div[3]/app-dash-cards/div/div[1]/div[1]/mat-card/mat-card-header/div[2]/mat-card-title')
            a.click()
            print("\n 2 - Selecting 'Certifications' successfully")
            self.ws['C3'] = 'Certifications'
            self.ws['G3'] = 'Pass'

            self.wb.save(self.filename)

            time.sleep(3)
        
        except Exception as e:
            print("\n\nERROR IN Certifications >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 2 - Selecting 'Certifications' get fail")
            self.ws['C3'] = 'Certifications'
            self.ws['G3'] = 'Fail'
            self.wb.save(self.filename)

    def Certifications_Api(self):

        try:
            self.url = "https://fonts.googleapis.com/css?family=Poppins"
            response = requests.get(self.url)
            print(response.content)
            print(response.headers)
        except Exception as e:
            print("Try again")

    def tearDown(self):
        self.driver.quit()

if __name__ == '__main__':
    tb = Api()
    tb.setUp()
    tb.Certifications()
    tb.Certifications_Api()