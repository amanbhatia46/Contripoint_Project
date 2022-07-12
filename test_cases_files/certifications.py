'''
FUncitonal

'''
from json import load
import unittest
import time
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta


class CAuth(unittest.TestCase):

    
    def __init__(self):
        self.filename = '/home/am.bhatia/Desktop/contripoint/ABC1.xlsx'
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'Certifications' not in self.wb.sheetnames:
            self.wb.create_sheet('Certifications')
            self.wb.save('/home/am.bhatia/Desktop/contripoint/ABC1.xlsx')
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'Certifications':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active            
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save('/home/am.bhatia/Desktop/contripoint/ABC1.xlsx')

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s=Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>> CERTIFICATIONS >>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            button = driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-login-screen/div/div/div/mat-card/div[2]/div/button')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="identifierId"]').send_keys('test.user@geminisolutions.in')
            self.ws['A2'] = 'test.user@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="identifierNext"]/div/button/span').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input').send_keys('Gemini@123#')
            self.ws['B2'] = 'Gemini@123#'
            driver.find_element(
                By.XPATH, '//*[@id="passwordNext"]/div/button/span').click()
            time.sleep(6)

            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

            
            


     # As per unittest module, individual test should start with test_
    def Certifications(self):
        """

        """
        try:
            #Certifications
            a = self.driver.find_element(By.XPATH,'/html/body/app-root/div[3]/div/app-dash-board/div[3]/app-dash-cards/div/div[1]/div[1]/mat-card/mat-card-header/div[2]/mat-card-title')
            a.click()
            print("\n 2 - Selecting 'Certifications' successfully")
            self.ws['C3'] = 'Certifications'
            self.ws['G3'] = 'Pass'

            self.wb.save(self.filename)

            time.sleep(3)
        
        except Exception as e:
            print("\n\nERROR IN Certifications >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 2 - Selecting 'Certifications' get fail")
            self.ws['C3'] = 'Certifications'
            self.ws['G3'] = 'Fail'
            self.wb.save(self.filename)

    def Add_New_Contribution(self):
        """

        """
        try:
            #Adding New Contribution
            self.driver.find_element(By.XPATH,'//*[@id="add_btn"]').click()
            time.sleep(6)
            print("\n 3 - Add New Contribution button gets selected")
            self.ws['C4'] = 'Add New Contribution'
            self.ws['G4'] = 'Pass'

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Adding New Contribution >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 3 - Add New Contribution button gets fail")
            self.ws['C4'] = 'Add New Contribution'
            self.ws['G4'] = 'Fail'

            self.wb.save(self.filename)

    def Name_Of_Certification(self):
        """

        """
        try:
            #NAME OF CERTIFICATION-
            self.driver.find_element(By.XPATH,'//*[@id="mat-input-0"]') .send_keys('Python')           
            print("\n 4 - Name Of Certification - 'Python'")
            self.ws['C5'] = 'Name Of Certification'
            self.ws['G5'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN Name Of Certificate >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 4 - Unable to select Name Of Certification from dropdown list ")
            self.ws['C5'] = 'Name Of Certification'
            self.ws['G5'] = 'Fail'

            self.wb.save(self.filename)

            
    def Date_Of_Completion(self):
        """

        """
        try:
            #Selecting Date of completion-
            datefield = self.driver.find_element(By.XPATH,'//*[@id="mat-dialog-0"]/app-certificate-add-modal/div/mat-dialog-content/form/div/div[2]/div/mat-form-field/div/div[1]/div[2]/mat-datepicker-toggle/button').click()
            datefield = self.driver.find_element(By.XPATH,"//div[@class='mat-calendar-body-cell-content mat-focus-indicator mat-calendar-body-today']").click()
            print("\n 5 - Date of completion done")
            self.ws['C6'] = 'Date of completion'
            self.ws['G6'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN Date of Completion >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 5 - Date of completion gets fail")
            self.ws['C6'] = 'Date of completion'
            self.ws['G6'] = 'Fail'

            self.wb.save(self.filename)

    def Uploading_Attachment(self):
        """

        """
        try:
            #Uploading Attatchment-
            self.driver.find_element(By.XPATH,'/html/body/div[2]/div[2]/div/mat-dialog-container/app-certificate-add-modal/div/mat-dialog-content/form/div/div[3]/div[2]/label').click()
            time.sleep(8)
            pyautogui.write('/home/am.bhatia/Desktop/image.jpeg') 
            pyautogui.press('enter')
            print("\n 6 - Attachment Upload successfully")
            self.ws['C7'] = 'Upload attatchment'
            self.ws['G7'] = 'Fail'

            self.wb.save(self.filename)
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN Uploading Attatchment >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 6 - Unable to upload the attatchment")
            self.ws['C7'] = 'Upload attatchment'
            self.ws['G7'] = 'Fail'

            self.wb.save(self.filename)

    def Technology(self):
        """

        """
        try:
            #Technology-
            self.driver.find_element(By.XPATH,'/html/body/div[2]/div[2]/div/mat-dialog-container/app-certificate-add-modal/div/mat-dialog-content/form/div/div[2]/div[2]/mat-form-field/div/div[1]/div').click()
            print("\n 7 - Selecting Technology. . .")
            self.ws['C8'] = 'Date of completion'
            self.ws['G8'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(3)

            self.driver.find_element(By.XPATH,'/html/body/div[2]/div[4]/div/div/div/mat-option[1]/mat-pseudo-checkbox').click()
            print("\n 8 - 'ANDROID' got selected from Technology")
            
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN TECHNOLOGY >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 7 - Unable to select Technology from dropdown list")
            self.ws['C8'] = 'Technology'
            self.ws['G8'] = 'Fail'

            self.wb.save(self.filename)


    def Submit(self):
        """

        """
        try:
            #SUBMIT
            element = self.driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-certificate-add-modal/div/div[2]/div[1]/button')
            actions = ActionChains(self.driver)
            actions.move_to_element(element).click().perform()

            self.driver.find_element(
                By.CSS_SELECTOR, ".col-xs-3 > #add_btn").click()
            print("\n 9 - All Details get SUBMIT successfully.")
            self.ws['C9'] = 'Submit'
            self.ws['G9'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

            #OK Button
            assert self.driver.find_element(
                By.CSS_SELECTOR, "#ok_btn > .mat-button-wrapper").text == "OK"
            time.sleep(5)

            element = self.driver.find_element(By.ID, "ok_btn")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            time.sleep(5)

            self.driver.find_element(By.ID, "ok_btn").click()
            print("\n 10 - Clicking OK Button")
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN SUBMIT  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 9 - Unable to Click Submit Button")
            self.ws['C9'] = 'Submit'
            self.ws['G9'] = 'Fail'

            self.wb.save(self.filename)
            time.sleep(3)


    def OK_Button(self):
        """
        
        """
        try:
            #OK Button
            assert self.driver.find_element(
                By.CSS_SELECTOR, "#ok_btn > .mat-button-wrapper").text == "OK"
            time.sleep(5)

            element = self.driver.find_element(By.ID, "ok_btn")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            time.sleep(5)

            self.driver.find_element(By.ID, "ok_btn").click()
            print("\n 11 - Clicking OK Button")
            self.ws['C10'] = 'OK Button'
            self.ws['G10'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN OK Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 11 - Unable to Click OK Button")
            self.ws['C10'] = 'OK Button'
            self.ws['G10'] = 'Fail'

            self.wb.save(self.filename)
    

    # Anything declared in tearDown will be executed for all test cases
    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = CAuth()
    tb.setUp()
    tb.Certifications()
    tb.Add_New_Contribution()
    tb.Name_Of_Certification()
    tb.Date_Of_Completion()
    tb.Uploading_Attachment()
    tb.Technology()
    tb.Submit()
    tb.OK_Button()
    #unittest.main()
    #unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))

