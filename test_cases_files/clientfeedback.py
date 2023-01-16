'''
FUncitonal

'''
import unittest
import time
import pyautogui
import HtmlTestRunner as x
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from json import load
from openpyxl import load_workbook
from datetime import datetime, timedelta

from datainputs import *

class CFAuth(unittest.TestCase):
    

    def __init__(self):
        
        """ interact with the underlying operating system."""
        import os
        print(os.path.join(os.getcwd()+ "\\xyz.xlsx"),">>>>>>")
        self.filename = os.path.join(os.getcwd()+ "\\xyz.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'ClientFeedback' not in self.wb.sheetnames:
            self.wb.create_sheet('ClientFeedback')
            self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'ClientFeedback':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))


        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s=Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>>> CLIENT FEEDBACK >>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            driver.implicitly_wait(10)  # seconds
            button = driver.find_element(
                By.XPATH, '//button[contains(., "Login with Gemini mail")]')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="i0116"]').send_keys(login_Id)
            self.ws['A2'] = 'aman.bhatia@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="idSIButton9"]').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="i0118"]').send_keys(login_Password)
            self.ws['B2'] = 'RitaNandini96'
            driver.find_element(
                By.XPATH, '//*[@id="idSIButton9"]').click()
            time.sleep(6)

            # Entering Multi-factor Authentication (MFA) Code Manually
            driver.find_element(By.ID, "idTxtBx_SAOTCC_OTC")
            time.sleep(6)
            print("\n Entering and Verifying MFA Code \n")

            driver.find_element(
                By.XPATH, '//*[@id="idSubmit_SAOTCC_Continue"]').click() 
            time.sleep(5)

            driver.find_element(
                By.XPATH, '//*[@id="KmsiCheckboxField"]').click()
            time.sleep(5)

            driver.find_element(By.ID, "idSIButton9").click()

            print("\n Login Successfull \n")


            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

                   
    def Clientfeedback(self):
        '''

        '''
        try:
            
            #ClientFeedback
            clientfeedback = self.driver.find_element(
                By.XPATH, '//*[contains(text(),"Client Feedback")]')
            clientfeedback.click()
            print("\n 2 - Selecting 'ClientFeedback'")
            self.ws['C3'] = 'Selecting ClientFeedback'
            self.ws['G3'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN CLIENT FEEDBACK >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 2 - Selecting 'ClientFeedback' gets failed")
            self.ws['C3'] = 'Selecting ClientFeedback'
            self.ws['G3'] = 'Fail'
            self.wb.save(self.filename)


    def AddNew(self):
        '''

        '''
        try:
            #ADD NEW
            element = self.driver.find_element(
                By.XPATH, '//*[@id="add_btn"]')
            actions = ActionChains(self.driver)
            actions.move_to_element(element).click().perform()
            
            print("\n 3 - Selecting 'Add New' button gets selected")
            self.ws['C4'] = 'Add New'
            self.ws['G4'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN ADD NEW  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 3 - Selecting 'Add New' button gets fail")
            self.ws['C4'] = 'Add New'
            self.ws['G4'] = 'Fail'
            self.wb.save(self.filename)


    def ProjectName(self):
        '''

        '''
        try:
            #ProjectName
            self.driver.find_element(
                By.XPATH, '//*[@id="mat-input-0"]').send_keys('Skedular')
            
            print("\n 4 - Project Name Done")
            self.ws['C5'] = 'Project Name'
            self.ws['G5'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN PROJECT NAME >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 4 - Unable to fill Project name")
            self.ws['C5'] = 'Project Name'
            self.ws['G5'] = 'Fail'
            self.wb.save(self.filename)


    def ClientName(self):
        '''

        '''
        try:
            #ClientName
            self.driver.find_element(
                By.ID, "mat-input-1").send_keys("auto-skedular")
            print("\n 5 - Client Name Done")
            self.ws['C6'] = 'Client Name'
            self.ws['G6'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN CLIENT NAME >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 5 - Unable to fill Client name")
            self.ws['C6'] = 'Client Name'
            self.ws['G6'] = 'Fail'
            self.wb.save(self.filename)

    def Enter_ClientFeedback(self):
        '''

        '''
        try:
            #Enter Client Feedback
            self.driver.find_element(
                By.ID, "mat-input-2").send_keys("US client")
            print("\n 6 - Client Feedback Done")
            self.ws['C7'] = 'Enter Client Feedback'
            self.ws['G7'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN CLIENT FEEDBACK >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 6 - Unable to fill Client Feedback")
            self.ws['C7'] = 'Enter Client Feedback'
            self.ws['G7'] = 'Fail'
            self.wb.save(self.filename)


    def Upload_attatchment(self):
        '''

        '''
        try:
            #Upload Attachment
            assert self.driver.find_element(
                By.ID, "attachment_btn").text == "Upload attachment"

            print("\n 7 - Uploading Attatchment . . . ")
            self.wb.save(self.filename)
            time.sleep(5)

            self.driver.find_element(By.ID, "attachment_btn").click()
            print("\n 8 - Selecting Attachment from Drive . . .")
            time.sleep(5)

            # self.driver.find_element(By.ID, "firsting").send_keys(
            #     'C:\\Users\\Aman Bhatia\\OneDrive - Gemini Solutions\\Desktop\\wallpaper\\1.jpeg')
            # print("\n 9 - Attatchment Successfully upload")
            # self.ws['C8'] = 'Upload attatchment'
            # self.ws['G8'] = 'Pass'
            # self.wb.save(self.filename)
            # time.sleep(5)

            pyautogui.write("C:\\Users\\Aman Bhatia\\OneDrive - Gemini Solutions\\Desktop\\wallpaper\\1.jpeg")
            pyautogui.press('enter') 
            print("\n 6 - Attachment Upload successfully")
            self.ws['C7'] = 'Upload attatchment'
            self.ws['G7'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN Uploading Attatchment >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 6 - Unable to upload the attatchment")
            self.ws['C7'] = 'Upload attatchment'
            self.ws['G7'] = 'Fail'

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN UPLOAD ATTACHMENT >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 9 - Unable to upload attatchment")
            self.ws['C8'] = 'Upload attatchment'
            self.ws['G8'] = 'Fail'
            self.wb.save(self.filename)

    def Goal_Type(self):
        """

        """
        try:
            #Goal Type-
            self.driver.find_element(By.XPATH,'//*[text()="Engineering Council (EC)"]').click()
            print("\n 10 - Selecting Goal Type- Engineering Council (EC)")
            self.ws['C9'] = 'Engineering Council (EC)'
            self.ws['G9'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(3)

        except Exception as e:
            print("\n\nERROR IN Goal Type >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 10 - Unable to select Goal_Type")
            self.ws['C9'] = 'Engineering Council (EC)'
            self.ws['G9'] = 'Fail'

            self.wb.save(self.filename)


    def Submit(self):
        '''

        '''
        try:
            #SUBMIT
            assert self.driver.find_element(
                By.CSS_SELECTOR, ".col-xs-3 > #add_btn > .mat-button-wrapper").text == "SUBMIT"

            self.driver.find_element(
                By.CSS_SELECTOR, ".col-xs-3 > #add_btn").click()

            print("\n 11 - All Details get SUBMIT successfully.")
            self.ws['C10'] = 'Submit'
            self.ws['G10'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN SUBMIT >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 11 - Unable to submit all details")
            self.ws['C10'] = 'Submit'
            self.ws['G10'] = 'Fail'
            self.wb.save(self.filename)



    def OK_Button(self):
        '''

        '''
        try:
            #OK Button
            assert self.driver.find_element(
                By.CSS_SELECTOR, "#ok_btn > .mat-button-wrapper").text == "OK"
            time.sleep(5)

            element = self.driver.find_element(By.ID, "ok_btn")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            time.sleep(5)

            self.driver.find_element(By.ID, "ok_btn").click()
            print("\n 12 - Clicking OK Button")
            self.ws['C11'] = 'OK Button'
            self.ws['G11'] = 'Fail'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN OKButton >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 12 - Unable to Click OK Button")
            self.ws['C11'] = 'OK Button'
            self.ws['G11'] = 'Fail'
            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = CFAuth()
    tb.setUp()
    tb.Clientfeedback()
    tb.AddNew()
    tb.ProjectName()
    tb.ClientName()
    tb.Enter_ClientFeedback()
    tb.Upload_attatchment()
    tb.Goal_Type()
    tb.Submit()
    tb.OK_Button()
    
    # unittest.main()
    # unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))

