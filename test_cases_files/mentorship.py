'''
FUncitonal

'''
from json import load
import unittest
import time
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta

class MAuth(unittest.TestCase):
    
    def __init__(self):
        
        """ interact with the underlying operating system."""
        import os
        print(os.path.join(os.getcwd()+ "\\xyz.xlsx"),">>>>>>")
        self.filename = os.path.join(os.getcwd()+ "\\xyz.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'Mentorship' not in self.wb.sheetnames:
            self.wb.create_sheet('Mentorship')
            self.wb.save('/C:Users/Aman Bhatia/OneDrive - Gemini Solutions/Desktop/Contripoint_Project/xyz.xlsx')
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'Mentorship':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s=Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>> MENTORSHIP >>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            button = driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-login-screen/div/div/div/mat-card/div[2]/div/button')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="identifierId"]').send_keys('aman.bhatia@geminisolutions.in')
            self.ws['A2'] = 'aman.bhatia@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="identifierNext"]/div/button/span').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input').send_keys('AmanBhatia96')
            self.ws['B2'] = 'AmanBhatia96'
            driver.find_element(
                By.XPATH, '//*[@id="passwordNext"]/div/button/span').click()
            time.sleep(6)

            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

    def Mentorship(self):
        '''

        '''
        try:
            
            # Mentorship
            Module = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-dash-board/div[3]/app-dash-cards/div/div[2]/div[1]/mat-card/mat-card-header/div[2]/mat-card-title')
            Module.click()
            print("\n 2 - Selecting 'Mentorship'")
            self.ws['C3'] = 'Mentorship'
            self.ws['G3'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(6)
        
        except Exception as e:
            print(e)
            print("\n 2 - Mentorship failed")
            self.ws['C3'] = 'Mentorship'
            self.ws['G3'] = 'Fail'
            self.wb.save(self.filename)
            
    
    def Add_New(self):
        """

        """
        try:      
            # ADD NEW
            ActionChains(self.driver).move_to_element(self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-mentorship-table/div/div/div/div[2]/div/div/div/div[2]/ul/div/button')).click()
            time.sleep(6)
            
            assert self.driver.find_element(
                By.CSS_SELECTOR, ".mat-button-wrapper").text == "ADD NEW"
            time.sleep(6)

            element = self.driver.find_element(
                By.CSS_SELECTOR, ".mat-button-wrapper")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            print("\n 3 - 'Add New' button gets selected")

            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN Adding New Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 3 - Add New button gets fail")
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Fail'

            self.wb.save(self.filename)

    def Colleagues_Mentored(self):
        """
        """
        try:
            #Colleagues Mentored
            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-button-wrapper").click()
            print("\n 4 - Selecting Mentors From The List . . .")
            time.sleep(6)

            assert self.driver.find_element(
                By.ID, "mat-select-value-1").text == "Colleagues Mentored *"
           
            time.sleep(6)

            self.driver.find_element(
                By.CSS_SELECTOR, ".ng-tns-c87-15:nth-child(2)").click()
            print("\n 5 - 'GSI G 960 Aaradhya Ahuja'")
            time.sleep(6)

            self.driver.find_element(
                By.CSS_SELECTOR, "#mat-option-35 > .mat-pseudo-checkbox").click()
            print("\n 6 - 'GSI G 1119 Aakriti Singhal'")
            time.sleep(6)

            self.driver.find_element(By.ID, "mat-option-36").click()
            time.sleep(6)

            self.driver.find_element(
                By.CSS_SELECTOR, ".cdk-overlay-transparent-backdrop").click()
            print("\n 8 - Mentors got selected")
            self.ws['C5'] = 'Mentorship'
            self.ws['G5'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN MENTORSHIP >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 8 - Selecting Mentors from dropdown list gets fail")
            self.ws['C5'] = 'Mentorship'
            self.ws['G5'] = 'Fail'

            self.wb.save(self.filename)

    def Mentorship_Duration(self):
        """

        """
        try:
            #Mentorship Duration
            self.driver.find_element(By.ID, "mat-input-1").send_keys("2")
            print("\n 9 - Duration in weeks - '2'")
            self.ws['C6'] = 'Mentorship'
            self.ws['G6'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN Mentorship Duration >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 9 - Selecting Mentors from dropdown list gets fail")
            self.ws['C6'] = 'Mentorship'
            self.ws['G6'] = 'Fail'

            self.wb.save(self.filename)

    def Enter_Description(self):
        """
        """
        try:

            #Enter Description
            self.driver.find_element(
                By.ID, "mat-input-2").send_keys("good work")
            print("\n 10 - Entering Description")
            self.ws['C7'] = 'Enter Description'
            self.ws['G7'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN Enter Description >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 10 - Entering Description gets fail")
            self.ws['C7'] = 'Enter Description'
            self.ws['G7'] = 'Fail'
            self.wb.save(self.filename)

    def Technology(self):
        """
        """
        try:    
            #Technology
            assert self.driver.find_element(
                By.ID, "mat-select-value-3").text == "Select Technology"
            print("\n 11 - Selecting 'Technology' from the list . . .")
            time.sleep(6)
            
            self.driver.find_element(
                By.CSS_SELECTOR, ".ng-tns-c87-17:nth-child(2)").click()
            print("\n 12 - 'Automation Testing'")
            time.sleep(6)
            
            self.driver.find_element(
                By.CSS_SELECTOR, "#mat-option-4 > .mat-pseudo-checkbox").click()
            print("\n 13 - 'ASP'")
            time.sleep(6)
            self.driver.find_element(
                By.CSS_SELECTOR, "#mat-option-3 > .mat-pseudo-checkbox").click()
            print("\n 14 - 'Technology' gets select")
            self.ws['C8'] = 'Technology'
            self.ws['G8'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN Technology >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 14 -'Technology' gets fail")
            self.ws['C8'] = 'Technology'
            self.ws['G8'] = 'Fail'
            self.wb.save(self.filename)

    def Submit(self):
        """
        
        """
        try:
            #SUBMIT
            assert self.driver.find_element(By.CSS_SELECTOR, ".col-xs-3 > #add_btn").text == "SUBMIT"
            print("\n 15 - Finding 'SUBMIT' button")
            time.sleep(5)

            Submit = ActionChains(self.driver).move_to_element(self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-mentorship-modal/div/div[2]/div[1]/button'))
            Submit.click().perform()
            time.sleep(5)
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-mentorship-modal/div/div[2]/div[1]/button').click()
            print("\n 16 - All Details get SUBMIT successfully.")

            self.ws['C9'] = 'Submit'
            self.ws['G9'] = 'Pass'
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN SUBMIT  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 16 - Unable to Click OK Button")
            self.ws['C9'] = 'Submit'
            self.ws['G9'] = 'Fail'

            self.wb.save(self.filename)
            time.sleep(3)

    def OK_Button(self):
        """
        
        """
        try:
            # OK Button
            assert self.driver.find_element(By.ID, "ok_btn").text == "OK"
            time.sleep(6)

            element = self.driver.find_element(By.ID, "ok_btn")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            time.sleep(6)

            self.driver.find_element(By.ID, "ok_btn").click()
            print("\n 17 - Clicking OK Button")
            self.ws['C10'] = 'OK Button'
            self.ws['G10'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(6)

        except Exception as e:
            print("\n\nERROR IN OK Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 17 - Unable to Click OK Button")
            self.ws['C10'] = 'OK Button'
            self.ws['G10'] = 'Fail'

            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = MAuth()
    tb.setUp()
    tb.Mentorship()
    tb.Add_New()
    tb.Colleagues_Mentored()
    tb.Mentorship_Duration()
    tb.Enter_Description()
    tb.Technology()
    tb.Submit()
    tb.OK_Button()
    
    # unittest.main()
#    unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))
