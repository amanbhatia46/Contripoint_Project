from json import load
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta


class PAuth(unittest.TestCase):

    def __init__(self):
        
        """ interact with the underlying operating system."""
        import os
        print(os.path.join(os.getcwd()+ "\\xyz.xlsx"),">>>>>>")
        self.filename = os.path.join(os.getcwd()+ "\\xyz.xlsx")
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'Projects' not in self.wb.sheetnames:
            self.wb.create_sheet('Projects')
            self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'Projects':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save(os.path.join(os.getcwd()+ "\\xyz.xlsx"))

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>> PROJECTS >>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            button = driver.find_element(
                By.XPATH, '//button[contains(., "Login with Gemini mail")]')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="identifierId"]').send_keys('test.user@geminisolutions.in')
            self.ws['A2'] = 'test.user@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="identifierNext"]/div/button/span').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input').send_keys('Gemini@123#')
            self.ws['B2'] = 'Gemini@123#'
            driver.find_element(
                By.XPATH, '//*[@id="passwordNext"]/div/button/span').click()
            time.sleep(6)

            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

    def Projects(self):
        """

        """
        try:
            # Projects -
            self.driver.find_element(
                By.CSS_SELECTOR, ".row:nth-child(2) > .col-lg-4:nth-child(2) .mat-card-title").click()
            print("\n 2 - Clicking on module - 'Projects'")
            self.ws['C3'] = 'Project'
            self.ws['G3'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN PROJECTS >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 2 - Selecting Proejct gets fail")
            self.ws['C3'] = 'Project'
            self.ws['G3'] = 'Fail'

            self.wb.save(self.filename)

            print("\n 3 - sample test case of entering 'PROJECTS' successfully open")

    def Add_New(self):
        """

        """
        try:
            # ADD Button -
            self.driver.find_element(By.XPATH, '//*[@id="add_btn"]').click()
            print("\n 4 - 'Add New' button gets selected")

            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Adding New Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 4 - Add New button gets fail")
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Fail'

            self.wb.save(self.filename)

    def Project_Type(self):
        """

        """
        try:
            # Select Project Type *
            self.driver.find_element(By.ID, "mat-input-0").click()
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-project-modal/div/mat-dialog-content/form/div/div[1]/div/div/div[2]/mat-radio-group/mat-radio-button[1]').click()
            print("\n 5 - Project Type gets selected as 'Inhouse'")

            self.ws['C5'] = 'Project type'
            self.ws['G5'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Project Type >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 6 - Add New Project Type gets fail")
            self.ws['C5'] = 'Project Type'
            self.ws['G5'] = 'Fail'

            self.wb.save(self.filename)

    def Project_Title(self):
        """
        """
        try:
            # Project Title -
            self.driver.find_element(
                By.ID, "mat-input-0").send_keys("contripoint")

            print("\n 7 - Project Title savesuccessfully")

            self.ws['C6'] = 'Project Title'
            self.ws['G6'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Project Title >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 7 - Project Title save unsuccessfully")
            self.ws['C6'] = 'Project Title'
            self.ws['G6'] = 'Fail'

            self.wb.save(self.filename)

    def Date_of_Completion(self):
        """

        """
        try:
            self.driver.find_element(
                By.CSS_SELECTOR, ".ng-tns-c85-17 > .mat-form-field-infix").click()
            time.sleep(4)  # calander
            print("\n 8 - Selecting Date Of Completion . . . ")

            element = self.driver.find_element(
                By.CSS_SELECTOR, ".mat-datepicker-toggle-default-icon")
            time.sleep(3)

            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()

            print("\n 9 - Date Of Completion get selected")
            time.sleep(3)

            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-datepicker-toggle-default-icon").click()
            time.sleep(3)
            print("\n 10 - Opening the calander . . .")

            date = self.driver.find_element(
                By.XPATH, '//*[@id="mat-datepicker-0"]/div/mat-month-view/table/tbody/tr[5]/td[6]/div[1]')
            date.click()
            time.sleep(5)
            print("\n 11 - Date Of Completion")
            self.ws['C7'] = 'Date Of Completion'
            self.ws['G7'] = 'Pass'

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Date Of Completion >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 11 - Selecting Date Of Completion gets fail")
            self.ws['C7'] = 'Date Of Completion'
            self.ws['G7'] = 'Fail'

            self.wb.save(self.filename)

    def Project_Status(self):
        """
        """
        try:
            ActionChains(self.driver).move_to_element(self.driver.find_element(
                By.CSS_SELECTOR, ".ng-tns-c85-19 > .mat-form-field-infix")).click()
            print("\n 12 - Selecting Project Status . . .")

            self.driver.find_element_by_xpath(
                '/html/body/div[2]/div[2]/div/mat-dialog-container/app-project-modal/div/mat-dialog-content/form/div/div[3]/div/div/div[2]/mat-radio-group/mat-radio-button[2]').click()
            print("\n 13 - Project Status Done")
            self.ws['C8'] = 'Project Status'
            self.ws['G8'] = 'Pass'
            time.sleep(4)

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Project Status >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 13 - Entering Project Status gets fail")
            self.ws['C8'] = 'Project Status'
            self.ws['G8'] = 'Fail'
            self.wb.save(self.filename)

    def Description(self):
        """

        """
        try:
            self.driver.find_element(
                By.ID, "mat-input-3").send_keys("internal project")
            time.sleep(4)
            print("\n 14 - Entering Description Done")
            self.ws['C8'] = 'Description'
            self.ws['G8'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Project Status >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 14 - Entering Description gets fail")
            self.ws['C8'] = 'Description'
            self.ws['G8'] = 'Fail'
            self.wb.save(self.filename)

    def Submit(self):
        """

        """
        try:
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-project-modal/div/div[2]/div[1]/button').click()
            time.sleep(4)
            print("\n 15 - All Details get submit successfully.")
            self.ws['C9'] = 'Submit'
            self.ws['G9'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN SUBMIT  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 15 - Unable to Click Submit Button")
            self.ws['C9'] = 'Submit'
            self.ws['G9'] = 'Fail'

            self.wb.save(self.filename)
            time.sleep(3)

    def OK_Button(self):
        """

        """
        try:
            self.driver.find_element(
                By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/app-confirmation-modal/div/div[4]/button').click()
            print("\n 16 - Clicking OK Button")
            self.ws['C10'] = 'OK Button'
            self.ws['G10'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN OK Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 16 - Unable to Click OK Button")
            self.ws['C10'] = 'OK Button'
            self.ws['G10'] = 'Fail'

            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = PAuth()
    tb.setUp()
    tb.Projects()
    tb.Add_New()
    tb.Project_Type()
    tb.Project_Title()
    tb.Date_of_Completion()
    tb.Project_Status()
    tb.Description()
    tb.Submit()
    tb.OK_Button()
    # unittest.main()
#    unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))
