from json import load
import unittest
import time
import HtmlTestRunner as x
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta


class SDAuth(unittest.TestCase):

    def __init__(self):
        self.filename = '/home/am.bhatia/Desktop/contripoint/abc1.xlsx'
        self.wb = load_workbook(filename=self.filename)
        self.index_sheet = 0
        if 'SelfDevelopment' not in self.wb.sheetnames:
            self.wb.create_sheet('SelfDevelopment')
            self.wb.save('/home/am.bhatia/Desktop/contripoint/abc1.xlsx')
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active
        for i, n in enumerate(self.wb.sheetnames):
            if n == 'SelfDevelopment':
                self.index_sheet = i
                break
        self.wb.active = self.index_sheet
        self.ws = self.wb.active
        self.wb.active = 1
        self.ws['A1'] = 'Email'
        self.ws['B1'] = 'Password'
        self.ws['C1'] = 'Test Case Module'
        self.ws['G1'] = 'Result'
        self.ws['H1'] = 'Date and Time'
        self.wb.save('/home/am.bhatia/Desktop/contripoint/abc1.xlsx')

        # datetime object containing current date and time
        futuredate = str(datetime.now())
        print(futuredate)
        self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>> PROJECTS >>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            button = driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-login-screen/div/div/div/mat-card/div[2]/div/button')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="identifierId"]').send_keys('test.user@geminisolutions.in')
            self.ws['A2'] = 'test.user@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="identifierNext"]/div/button/span').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input').send_keys('Gemini@123#')
            self.ws['B2'] = 'Gemini@123#'
            driver.find_element(
                By.XPATH, '//*[@id="passwordNext"]/div/button/span').click()
            time.sleep(6)

            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)

    def SelfDevelopment(self):
        '''

        '''
        try:
            
            # SelfDevelopment
            SelfDevelopment = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-dash-board/div[3]/app-dash-cards/div/div[3]/div[2]/mat-card/mat-card-header')
            SelfDevelopment.click()
            print("\n 2 - Selecting 'SelfDevelopment'")
            self.ws['C3'] = 'SelfDevelopment'
            self.ws['G3'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN SelfDevelopment >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 2 - Selecting SelfDevelopment gets fail")
            self.ws['C3'] = 'SelfDevelopment'
            self.ws['G3'] = 'Fail'

            self.wb.save(self.filename)
            time.sleep(6)

    def Add_New(self):
        '''

        '''
        try:

            # ADD NEW
            element = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-self-development-table/div/div/div/div[2]/div/div/div/div[2]/ul/div/button')
            actions = ActionChains(self.driver)
            actions.move_to_element(element).click().perform()
            print("\n 3 - Selecting 'Add New' button gets selected")
            time.sleep(6)
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Pass'

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Adding New Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 3 - Add New button gets fail")
            self.ws['C4'] = 'Add New Button'
            self.ws['G4'] = 'Fail'

            self.wb.save(self.filename)


    def Skill_Name(self):
        '''

        '''
        try:

            # Name of skill developed
            assert self.driver.find_element(
                By.CSS_SELECTOR, ".ng-tns-c85-15 > .mat-form-field-infix").text == "Name of skill developed *"
            time.sleep(5)

            self.driver.find_element(By.ID, "mat-input-0").click()
            time.sleep(5)

            self.driver.find_element(
                By.ID, "mat-input-0").send_keys("Machine Learning")
            print("\n 4 - Name of Skill - 'Machine Learning'")
            time.sleep(5)
            self.ws['C5'] = 'Skill_Name'
            self.ws['G5'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Skill Name >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 4 - Adding Skill_Name gets fail")
            self.ws['C5'] = 'Skill_Name'
            self.ws['G5'] = 'Fail'

            self.wb.save(self.filename)


    def Duration(self):
        '''

        '''
        try:

            # Duration
            self.driver.find_element(By.ID, "mat-input-1").click()
            time.sleep(5)

            self.driver.find_element(By.ID, "mat-input-1").send_keys("2")
            print("\n 5 - Duration in weeks- '2'")
            self.ws['C6'] = 'Duration'
            self.ws['G6'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Duration >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 5 - Adding Duration gets fail")
            self.ws['C6'] = 'Duration'
            self.ws['G6'] = 'Fail'

            self.wb.save(self.filename)
            

    def StartDate(self):
        '''

        '''
        try:

            # Start Date
            self.driver.find_element(By.ID, "mat-input-2").click()
            print("\n 6 - Opening Calander . . .")
            time.sleep(5)

            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-calendar-body-today").click()
            print("\n 7 - Date Done")
            time.sleep(5)
            self.ws['C6'] = 'Start Date'
            self.ws['G6'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN Start Date >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 7 - Start Date gets fail")
            self.ws['C6'] = 'Start Date'
            self.ws['G6'] = 'Fail'

            self.wb.save(self.filename)

    def Description(self):
        """

        """
        try:
            #Enter Description
            self.driver.find_element(By.ID, "mat-input-3").click()
            time.sleep(5)

            self.driver.find_element(By.ID, "mat-input-3").send_keys(
                "Machine learning is the study of computer algorithms that can improve automatically through experience and by the use of data. It is seen as a part of artificial intelligence")
            print("\n 8 - Entering Description Done")
            time.sleep(5)
            self.ws['C7'] = 'Description'
            self.ws['G7'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN Project Status >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 8 - Entering Description gets fail")
            self.ws['C7'] = 'Description'
            self.ws['G7'] = 'Fail'
            self.wb.save(self.filename)

    def Submit(self):
        """

        """
        try:

            #SUBMIT
            assert self.driver.find_element(
                By.CSS_SELECTOR, ".col-xs-3 > #add_btn").text == "SUBMIT"
            time.sleep(5)

            self.driver.find_element(
                By.CSS_SELECTOR, ".col-xs-3 > #add_btn").click()
            time.sleep(5)

            print("\n 9 - All Details get SUBMIT successfully.")
            self.ws['C8'] = 'Submit'
            self.ws['G8'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN SUBMIT  >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 9 - Unable to Click Submit Button")
            self.ws['C8'] = 'Submit'
            self.ws['G8'] = 'Fail'

            self.wb.save(self.filename)

    def OK_Button(self):
        """

        """
        try:

            #OK Button
            assert self.driver.find_element(By.ID, "ok_btn").text == "OK"
            time.sleep(5)

            element = self.driver.find_element(By.ID, "ok_btn")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            time.sleep(5)

            self.driver.find_element(By.ID, "ok_btn").click()
            print("\n 10 - Clicking OK Button")
            self.ws['C9'] = 'OK Button'
            self.ws['G9'] = 'Pass'

            self.wb.save(self.filename)
            time.sleep(4)

        except Exception as e:
            print("\n\nERROR IN OK Button >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 10 - Unable to Click OK Button")
            self.ws['C9'] = 'OK Button'
            self.ws['G9'] = 'Fail'

            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()


if __name__ == '__main__':
    tb = SDAuth()
    tb.setUp()
    tb.SelfDevelopment()
    tb.Add_New()
    tb.Skill_Name()
    tb.Duration()
    tb.StartDate()
    tb.Description()
    tb.Submit()
    tb.OK_Button()
    # unittest.main()
#    unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))

