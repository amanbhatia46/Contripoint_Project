'''
FUncitonal

'''
from json import load
import unittest
import time
import HtmlTestRunner as x
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook
from datetime import datetime, timedelta



class TBAuth(unittest.TestCase):
    

    def __init__(self):
        def __init__(self):
            self.filename = '/home/am.bhatia/Desktop/contripoint/ABC1.xlsx'
            self.wb = load_workbook(filename=self.filename)
            self.index_sheet = 0
            if 'TeamBuildingActivity' not in self.wb.sheetnames:
                self.wb.create_sheet('TeamBuildingActivity')
                self.wb.save('/home/am.bhatia/Desktop/contripoint/ABC1.xlsx')
            self.wb = load_workbook(filename=self.filename)
            self.ws = self.wb.active
            for i, n in enumerate(self.wb.sheetnames):
                if n == 'TeamBuildingActivity':
                    self.index_sheet = i
                    break
            self.wb.active = self.index_sheet
            self.ws = self.wb.active
            self.wb.active = 1
            self.ws['A1'] = 'Email'
            self.ws['B1'] = 'Password'
            self.ws['C1'] = 'Test Case Module'
            self.ws['G1'] = 'Result'
            self.ws['H1'] = 'Date and Time'
            self.wb.save('/home/am.bhatia/Desktop/contripoint/ABC1.xlsx')

            # datetime object containing current date and time
            futuredate = str(datetime.now())
            print(futuredate)
            self.ws['H2'] = futuredate

    def setUp(self):
        try:
            s=Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=s)
            print("\n\n\n\n\n>>>>>>>>>>TEAM BUILDING ACTIVITY>>>>>>>>>>>>")

            driver = self.driver
            driver.maximize_window()
            driver.get(
                "https://dev-contripoint.geminisolutions.com/#/login")
            button = driver.find_element(
                By.XPATH, '/html/body/app-root/div/div/app-login-screen/div/div/div/mat-card/div[2]/div/button')
            button.click()
            time.sleep(6)

            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[1])
            driver.find_element(
                By.XPATH, '//*[@id="identifierId"]').send_keys('aman.bhatia@geminisolutions.in')
            self.ws['A2'] = 'aman.bhatia@geminisolutions.in'
            driver.find_element(
                By.XPATH, '//*[@id="identifierNext"]/div/button/span').click()
            time.sleep(3)

            driver.find_element(
                By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input').send_keys('AmanBhatia96')
            self.ws['B2'] = 'AmanBhatia96'
            driver.find_element(
                By.XPATH, '//*[@id="passwordNext"]/div/button/span').click()
            time.sleep(6)

            driver.switch_to.window(window_handles[0])
            time.sleep(6)
            print("\n 1 - Gemini Id and Password successfully login")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login Success'

            self.wb.save(self.filename)

        except Exception as e:
            print(e)
            print("\n 1 - Gemini Id and Password  login failed")
            self.ws['C2'] = 'Login'
            self.ws['G2'] = 'login failed'
            self.wb.save(self.filename)
       
    def teambuildingactivity(self):
        '''

        '''
        try:
            #TeamBuildingActivity
            TeamBuildingActivity = self.driver.find_element(
                By.XPATH, '/html/body/app-root/div[3]/div/app-dash-board/div[3]/app-dash-cards/div/div[2]/div[3]/mat-card/mat-card-header')
            TeamBuildingActivity.click()
            print("\n 2 - Selecting 'Team Building Activity'")
            self.ws['C3'] = 'Team Building Activity'
            self.ws['G3'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN TEAM BUILDING ACTIVITY >>>>>>>>>>>>>>>\n\n")
            print (e)
            print("\n 2 - Selecting 'Team Building Activity' got Failed")
            self.ws['C3'] = 'Team Building Activity'
            self.ws['G3'] = 'Fail'
            self.wb.save(self.filename)

        
    def AddNew(self):
        '''

        '''
        try:
            #ADD NEW
            assert self.driver.find_element(By.ID, "add_btn").text == "ADD NEW"
            time.sleep(5)
            
            self.driver.find_element(By.ID, "add_btn").click()
            print("\n 3 - 'Add New' button gets selected")
            time.sleep(5)

            self.ws['C4'] = 'Add New button'
            self.ws['G4'] = 'Pass'
            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN AddNew >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 3 - 'Add New' button got fail")
            self.ws['C4'] = 'Add New button'
            self.ws['G4'] = 'Fail'
            self.wb.save(self.filename)

    def NameofActivityConducted(self):
        '''

        '''
        try:
            # Name of Activity Conducted
            self.driver.find_element(By.ID, "mat-input-0").click()
            time.sleep(5)

            self.driver.find_element(
                By.ID, "mat-input-0").send_keys("Project")
            print("\n 4 - Name of Activity Conducted")

            self.ws['C5'] = 'Name of Activity Conducted'
            self.ws['G5'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN NameOfActivityConducted >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 3 - Name of Activity Fail")
            self.ws['C5'] = 'Name of Activity Conducted'
            self.ws['G5'] = 'Fail'
            self.wb.save(self.filename)


    def NumberOfParticipants(self):
        '''

        '''
        try:
            #No. Of Participants in the Activity
            self.driver.find_element(By.ID, "mat-input-2").send_keys("2")
            print("\n 5 - No. Of Participants in the Activity - '2'")

            self.ws['C6'] = 'No. Of Participants in the Activity'
            self.ws['G6'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN NumberOfParticipants >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 5 - No. Of Participants in the Activity gets fail")
            self.ws['C6'] = 'No. Of Participants in the Activity'
            self.ws['G6'] = 'Fail'
            self.wb.save(self.filename)
            

    def EnterDescription(self):
        '''

        '''
        try:
            #Enter Description 
            self.driver.find_element(By.ID, "mat-input-3").send_keys("If you want to lift yourself up, lift up with the team")
            print("\n 6 - Entering Description Done")

            self.ws['C7'] = 'Description'
            self.ws['G7'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

        except Exception as e:
            print("\n\nERROR IN EnterDescription >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 5 - Entering Description gets fail")
            self.ws['C7'] = 'Description'
            self.ws['G7'] = 'Fail'
            self.wb.save(self.filename)
        
    def DateoftheActivity(self):
        '''

        '''
        try:
            #Date of the Activity
            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-datepicker-toggle-default-icon").click()
            print("\n 7 - Selecting 'Date of the Activity' . . . ")

            self.ws['C8'] = 'Date of the Activity'
            self.ws['G8'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

            #Date from the Calander
            self.driver.find_element(
                By.CSS_SELECTOR, ".mat-calendar-body-today").click()
            print("\n 8 - Selecting the Date of the Activity from the calander")
            time.sleep(5)
        
        except Exception as e:
            print("\n\nERROR IN DateoftheActivity >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 7 - Fail to select the Date of the Activity from the calander")
            self.ws['C8'] = 'Date of the Activity'
            self.ws['G8'] = 'Fail'
            self.wb.save(self.filename)

    def SubmitButton(self):
        '''

        '''
        try:
            #SUBMIT Button
            assert self.driver.find_element(
                By.CSS_SELECTOR, ".col-xs-3 > #add_btn").text == "ADD"
            time.sleep(5)

            self.driver.find_element(
                By.CSS_SELECTOR, ".col-xs-3 > #add_btn").click()
            print("\n 8 - All Details get SUBMIT successfully.")

            self.ws['C9'] = 'SUBMIT'
            self.ws['G9'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)
        
        except Exception as e:
            print("\n\nERROR IN SubmitButton >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 8 - All Details get fail while SUBMIT.")
            self.ws['C9'] = 'SUBMIT'
            self.ws['G9'] = 'Fail'
            self.wb.save(self.filename)

    def OKButton(self):
        '''

        '''
        try:
            #OK Button
            assert self.driver.find_element(
                By.CSS_SELECTOR, "#ok_btn > .mat-button-wrapper").text == "OK"
            time.sleep(5)

            element = self.driver.find_element(By.ID, "ok_btn")
            actions = ActionChains(self.driver)
            actions.move_to_element(element).perform()
            time.sleep(5)

            self.driver.find_element(By.ID, "ok_btn").click()
            print("\n 9 - Clicking OK Button")

            self.ws['C10'] = 'OK BUTTON'
            self.ws['G10'] = 'Pass'
            self.wb.save(self.filename)
            time.sleep(5)

            self.wb.save(self.filename)

        except Exception as e:
            print("\n\nERROR IN OKButton >>>>>>>>>>>>>>>\n\n")
            print(e)
            print("\n 9 - Unable to click OK Button")
            self.ws['C10'] = 'OK BUTTON'
            self.ws['G10'] = 'Fail'
            self.wb.save(self.filename)

    def tearDown(self):
        self.driver.quit()
    

if __name__ == '__main__':
    tb= TBAuth()
    tb.setUp()
    tb.teambuildingactivity()
    tb.AddNew()
    tb.NameofActivityConducted()
    tb.NumberOfParticipants()
    tb.EnterDescription()
    tb.DateoftheActivity()
    tb.SubmitButton()
    tb.OKButton()

    # unittest.main()
    # unittest.main(testRunner= x.HTMLTestRunner( '/home/am.bhatia/Desktop/ContriPoint/testsuits'))

